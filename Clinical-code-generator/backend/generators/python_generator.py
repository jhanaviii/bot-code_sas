"""
Python Code Generator for Clinical Trial Submissions
Generates CDISC-compliant Python programs using pandas, statsmodels, and lifelines.
"""

from datetime import datetime


class PythonCodeGenerator:
    def __init__(self, config):
        self.config = config
        self.analysis_type = config.get('analysis_type', 'descriptive')
        self.output_type = config.get('output_type', 'table')
        self.input_data = config.get('input_data', {})
        self.output_data = config.get('output_data', {})
        self.variables = config.get('variables', {})
        self.analysis_params = config.get('analysis_parameters', {})
        self.submission = config.get('submission_details', {})

    def parse_shell(self, shell_path):
        metadata = {'columns': [], 'rows': [], 'titles': [], 'footnotes': []}
        return metadata

    def parse_specification(self, spec_path):
        metadata = {'variables': [], 'derivations': [], 'formats': []}
        return metadata

    def generate(self, shell_metadata=None, spec_metadata=None):
        sections = [
            self._generate_header(),
            self._generate_imports(),
            self._generate_setup(),
            self._generate_data_read(),
            self._generate_data_prep(),
            self._generate_analysis(),
            self._generate_output(),
            self._generate_validation()
        ]
        return '\n\n'.join(sections)

    def _generate_header(self):
        study_id = self.submission.get('study_id', 'STUDY-XXX')
        sponsor = self.submission.get('sponsor', 'AstraZeneca')
        protocol = self.submission.get('protocol', 'DXXXXCXXXXX')

        return f'''"""
================================================================================
PROGRAM NAME  : {self.output_data.get('dataset_name', 'program')}.py
STUDY         : {study_id}
PROTOCOL      : {protocol}
SPONSOR       : {sponsor}
PURPOSE       : {self._get_purpose_description()}
INPUT         : {', '.join(self.input_data.get('datasets', []))}
OUTPUT        : {self.output_data.get('dataset_name', 'output')}.{self.output_data.get('format', 'rtf')}
ANALYSIS TYPE : {self.analysis_type.replace('_', ' ').title()}
OUTPUT TYPE   : {self.output_type.title()}
POPULATION    : {self.analysis_params.get('population', 'safety').replace('_', ' ').title()}

AUTHOR        : [Programmer Name]
DATE CREATED  : {datetime.now().strftime('%d-%b-%Y')}
DATE MODIFIED : {datetime.now().strftime('%d-%b-%Y')}

REGULATORY    : {self.submission.get('regulatory_authority', 'FDA')}
SUBMISSION    : {self.submission.get('submission_type', 'NDA')}

VALIDATION    : Independent double programming required
CDISC VERSION : ADaM IG v1.3 / SDTM IG v3.4
================================================================================
"""'''

    def _generate_imports(self):
        base_imports = [
            'import pandas as pd',
            'import numpy as np',
            'from pathlib import Path',
            'from datetime import datetime',
            'import warnings',
            'warnings.filterwarnings("ignore")'
        ]

        analysis_imports = {
            'descriptive': [
                'from scipy import stats',
                'import statsmodels.api as sm'
            ],
            'efficacy': [
                'import statsmodels.api as sm',
                'from statsmodels.formula.api import ols, mixedlm',
                'from statsmodels.stats.multicomp import pairwise_tukeyhsd'
            ],
            'safety': [
                'from scipy import stats'
            ],
            'survival': [
                'from lifelines import KaplanMeierFitter, CoxPHFitter',
                'from lifelines.statistics import logrank_test',
                'import matplotlib.pyplot as plt'
            ],
            'mixed_model': [
                'import statsmodels.api as sm',
                'from statsmodels.formula.api import mixedlm',
                'from statsmodels.regression.mixed_linear_model import MixedLM'
            ],
            'categorical': [
                'from scipy.stats import chi2_contingency, fisher_exact',
                'import statsmodels.api as sm'
            ],
            'pk': [
                'from scipy import stats',
                'from scipy.integrate import trapezoid'
            ],
            'subgroup': [
                'import statsmodels.api as sm',
                'from statsmodels.formula.api import ols',
                'import matplotlib.pyplot as plt'
            ]
        }

        output_imports = {
            'table': ['from openpyxl import Workbook', 'from docx import Document'],
            'figure': ['import matplotlib.pyplot as plt', 'import seaborn as sns', 'from matplotlib.backends.backend_pdf import PdfPages'],
            'listing': ['from openpyxl import Workbook'],
            'dataset': ['import pyreadstat']
        }

        all_imports = base_imports + \
                      analysis_imports.get(self.analysis_type, []) + \
                      output_imports.get(self.output_type, [])

        # Deduplicate
        seen = set()
        unique_imports = []
        for imp in all_imports:
            if imp not in seen:
                seen.add(imp)
                unique_imports.append(imp)

        return '\n'.join(unique_imports)

    def _generate_setup(self):
        population_flag = self._get_population_flag()
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')

        return f"""# --- Program Configuration ---
CONFIG = {{
    'study_id': '{self.submission.get('study_id', 'STUDY001')}',
    'protocol': '{self.submission.get('protocol', 'D1234C00001')}',
    'prog_name': '{self.output_data.get('dataset_name', 'program')}',
    'output_name': '{self.output_data.get('dataset_name', 'output')}',
    'pop_flag': '{population_flag}',
    'trt_var': '{treatment_var}',
    'alpha': {self.analysis_params.get('alpha', 0.05)},
    'conf_level': {self.analysis_params.get('confidence_level', 0.95)},
    'adam_path': Path(r'{self.input_data.get('library_path', '/data/adam')}'),
    'output_path': Path(r'{self.output_data.get('output_path', '/output')}'),
}}

# Treatment ordering
TRT_ORDER = ['Placebo', 'Treatment A', 'Treatment B', 'Total']
TRT_MAP = {{'Placebo': 1, 'Treatment A': 2, 'Treatment B': 3, 'Total': 99}}

# Ensure output directory exists
CONFIG['output_path'].mkdir(parents=True, exist_ok=True)"""

    def _generate_data_read(self):
        datasets = self.input_data.get('datasets', ['adsl'])
        population_flag = self._get_population_flag()
        filter_vars = self.variables.get('filter_variables', [])

        read_blocks = []
        for ds in datasets:
            filters = [f'(df["{population_flag}"] == "Y")']
            for fv in filter_vars:
                if fv != population_flag:
                    filters.append(f'(df["{fv}"] == "Y")')

            filter_expr = ' & '.join(filters)

            read_blocks.append(f"""# Read {ds.upper()}
{ds}_path = CONFIG['adam_path'] / '{ds}.sas7bdat'
{ds}_raw, meta = pyreadstat.read_sas7bdat(str({ds}_path))
{ds} = {ds}_raw[{filter_expr}].copy()
{ds}[CONFIG['trt_var']] = pd.Categorical(
    {ds}[CONFIG['trt_var']], categories=TRT_ORDER[:-1], ordered=True
)
print(f"NOTE: {ds.upper()} has {{{ds}.shape[0]}} records after filtering.")""")

        return f"""# --- Read Input Datasets ---
import pyreadstat

{chr(10).join(read_blocks)}

# --- Get Population Counts (Big N) ---
big_n = {datasets[0]}.drop_duplicates(subset=['USUBJID']).groupby(
    CONFIG['trt_var']
)['USUBJID'].nunique().to_dict()

print(f"NOTE: Population counts: {{big_n}}")"""

    def _generate_data_prep(self):
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        sort_vars = self.variables.get('sort_variables', ['USUBJID'])

        return f"""# --- Data Preparation ---
analysis_data = {self.input_data.get('datasets', ['adsl'])[0]}.copy()

# Add treatment numeric ordering
analysis_data['TRTN'] = analysis_data[CONFIG['trt_var']].map(TRT_MAP)

# Sort data
analysis_data = analysis_data.sort_values(
    ['TRTN'] + {sort_vars}
).reset_index(drop=True)

# Create version with Total group
analysis_total = analysis_data.copy()
analysis_total[CONFIG['trt_var']] = 'Total'
analysis_total['TRTN'] = 99

analysis_with_total = pd.concat(
    [analysis_data, analysis_total], ignore_index=True
)

print(f"NOTE: Analysis dataset has {{analysis_with_total.shape[0]}} records (including Total).")"""

    def _generate_analysis(self):
        method_generators = {
            'descriptive': self._gen_descriptive_py,
            'efficacy': self._gen_efficacy_py,
            'safety': self._gen_safety_py,
            'survival': self._gen_survival_py,
            'mixed_model': self._gen_mmrm_py,
            'categorical': self._gen_categorical_py,
            'pk': self._gen_pk_py,
            'subgroup': self._gen_subgroup_py
        }
        generator = method_generators.get(self.analysis_type, self._gen_descriptive_py)
        return generator()

    def _gen_descriptive_py(self):
        analysis_vars = self.variables.get('analysis_variables', ['AVAL'])
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')

        return f"""# --- Descriptive Statistics ---

def compute_descriptive_stats(df, var, group_var):
    \"\"\"Compute descriptive statistics by treatment group.\"\"\"
    stats_df = df.groupby(group_var)[var].agg([
        ('n', 'count'),
        ('mean', 'mean'),
        ('std', 'std'),
        ('median', 'median'),
        ('min', 'min'),
        ('max', 'max'),
        ('q1', lambda x: x.quantile(0.25)),
        ('q3', lambda x: x.quantile(0.75))
    ]).reset_index()
    
    # Format for display
    stats_df['n_c'] = stats_df['n'].apply(lambda x: f"{{x:.0f}}")
    stats_df['mean_sd'] = stats_df.apply(
        lambda r: f"{{r['mean']:.1f}} ({{r['std']:.2f}})", axis=1
    )
    stats_df['median_c'] = stats_df['median'].apply(lambda x: f"{{x:.1f}}")
    stats_df['q1_q3'] = stats_df.apply(
        lambda r: f"{{r['q1']:.1f}}, {{r['q3']:.1f}}", axis=1
    )
    stats_df['min_max'] = stats_df.apply(
        lambda r: f"{{r['min']:.1f}}, {{r['max']:.1f}}", axis=1
    )
    
    return stats_df

# Compute statistics for each analysis variable
results = {{}}
for var in {analysis_vars}:
    results[var] = compute_descriptive_stats(
        analysis_with_total, var, CONFIG['trt_var']
    )
    print(f"\\nDescriptive Statistics for {{var}}:")
    print(results[var][['{ treatment_var}', 'n_c', 'mean_sd', 'median_c', 'q1_q3', 'min_max']])"""

    def _gen_efficacy_py(self):
        analysis_vars = self.variables.get('analysis_variables', ['CHG'])
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        covariates = self.analysis_params.get('covariates', ['BASE'])

        return f"""# --- Efficacy Analysis: ANCOVA ---

# Fit ANCOVA model
formula = '{analysis_vars[0]} ~ C({treatment_var}, Treatment("Placebo")) + {" + ".join(covariates)}'
model = ols(formula, data=analysis_data).fit()

print("\\nANCOVA Model Summary:")
print(model.summary())

# LS Means (adjusted means)
from statsmodels.stats.contrast import ContrastResults

# Get predicted means for each treatment
lsmeans = {{}}
for trt in analysis_data[CONFIG['trt_var']].unique():
    subset = analysis_data[analysis_data[CONFIG['trt_var']] == trt]
    pred_data = subset.copy()
    for cov in {covariates}:
        pred_data[cov] = analysis_data[cov].mean()
    lsmeans[trt] = model.predict(pred_data).mean()

print("\\nLS Means:")
for trt, mean in lsmeans.items():
    print(f"  {{trt}}: {{mean:.2f}}")

# Treatment differences
print("\\nTreatment Differences vs Placebo:")
for trt in [t for t in lsmeans.keys() if t != 'Placebo']:
    diff = lsmeans[trt] - lsmeans['Placebo']
    print(f"  {{trt}} - Placebo: {{diff:.2f}}")"""

    def _gen_safety_py(self):
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')

        return f"""# --- Safety Analysis: Adverse Event Summary ---

# Overall AE incidence
ae_overall = adae.groupby(CONFIG['trt_var']).agg(
    n_subjects=('USUBJID', 'nunique'),
    n_events=('USUBJID', 'count')
).reset_index()

ae_overall['N'] = ae_overall[CONFIG['trt_var']].map(big_n)
ae_overall['pct'] = (ae_overall['n_subjects'] / ae_overall['N']) * 100
ae_overall['col'] = ae_overall.apply(
    lambda r: f"{{r['n_subjects']:.0f}} ({{r['pct']:.1f}}%)", axis=1
)

print("\\nOverall AE Summary:")
print(ae_overall)

# AE by SOC and PT
ae_soc_pt = adae.groupby(
    [CONFIG['trt_var'], 'AEBODSYS', 'AEDECOD']
).agg(
    n_subjects=('USUBJID', 'nunique'),
    n_events=('USUBJID', 'count')
).reset_index()

ae_soc_pt['N'] = ae_soc_pt[CONFIG['trt_var']].map(big_n)
ae_soc_pt['pct'] = (ae_soc_pt['n_subjects'] / ae_soc_pt['N']) * 100
ae_soc_pt['col'] = ae_soc_pt.apply(
    lambda r: f"{{r['n_subjects']:.0f}} ({{r['pct']:.1f}}%)", axis=1
)

# Sort by descending frequency
ae_sorted = ae_soc_pt.sort_values(
    ['n_subjects', 'AEBODSYS', 'AEDECOD'], ascending=[False, True, True]
)

print("\\nTop 10 AEs by SOC/PT:")
print(ae_sorted.head(10))"""

    def _gen_survival_py(self):
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')

        return f"""# --- Survival Analysis ---

# Kaplan-Meier Estimation
kmf = KaplanMeierFitter()

fig, ax = plt.subplots(figsize=(10, 7))
km_results = {{}}

for trt in analysis_data[CONFIG['trt_var']].unique():
    subset = analysis_data[analysis_data[CONFIG['trt_var']] == trt]
    kmf.fit(
        durations=subset['AVAL'],
        event_observed=(1 - subset['CNSR']),
        label=trt
    )
    kmf.plot_survival_function(ax=ax, ci_show=True)
    
    km_results[trt] = {{
        'median': kmf.median_survival_time_,
        'survival_6mo': kmf.predict(180),
        'survival_12mo': kmf.predict(365)
    }}

ax.set_xlabel('Time (Days)')
ax.set_ylabel('Survival Probability')
ax.set_title('Kaplan-Meier Survival Curve')
ax.legend(loc='lower left')

# Log-rank test
groups = analysis_data[CONFIG['trt_var']].unique()
if len(groups) == 2:
    g1 = analysis_data[analysis_data[CONFIG['trt_var']] == groups[0]]
    g2 = analysis_data[analysis_data[CONFIG['trt_var']] == groups[1]]
    lr_result = logrank_test(
        g1['AVAL'], g2['AVAL'],
        event_observed_A=(1 - g1['CNSR']),
        event_observed_B=(1 - g2['CNSR'])
    )
    print(f"\\nLog-Rank Test p-value: {{lr_result.p_value:.4f}}")

# Cox Proportional Hazards
cph = CoxPHFitter()
cox_data = analysis_data[['AVAL', 'CNSR', CONFIG['trt_var']]].copy()
cox_data['event'] = 1 - cox_data['CNSR']
cox_data = pd.get_dummies(cox_data, columns=[CONFIG['trt_var']], drop_first=True)

cph.fit(cox_data, duration_col='AVAL', event_col='event')
print("\\nCox PH Model:")
cph.print_summary()"""

    def _gen_mmrm_py(self):
        analysis_vars = self.variables.get('analysis_variables', ['CHG'])
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        covariates = self.analysis_params.get('covariates', ['BASE'])

        return f"""# --- Mixed Model Repeated Measures (MMRM) ---

# Note: For full MMRM with unstructured covariance, consider using R via rpy2
# This implementation uses statsmodels mixed linear model

formula = '{analysis_vars[0]} ~ C({treatment_var}) * C(AVISIT) + {" + ".join(covariates)}'

mmrm_model = mixedlm(
    formula,
    data=analysis_data,
    groups='USUBJID',
    re_formula='~C(AVISIT)'
)

mmrm_result = mmrm_model.fit(reml=True)
print("\\nMMRM Model Summary:")
print(mmrm_result.summary())

# LS Means by visit
visits = analysis_data['AVISIT'].unique()
lsmeans_by_visit = {{}}

for visit in sorted(visits):
    visit_data = analysis_data[analysis_data['AVISIT'] == visit]
    for trt in visit_data[CONFIG['trt_var']].unique():
        trt_visit_data = visit_data[visit_data[CONFIG['trt_var']] == trt]
        pred = mmrm_result.predict(trt_visit_data)
        lsmeans_by_visit[(trt, visit)] = pred.mean()

print("\\nLS Means by Treatment and Visit:")
for (trt, visit), mean in sorted(lsmeans_by_visit.items()):
    print(f"  {{trt}} - {{visit}}: {{mean:.2f}}")"""

    def _gen_categorical_py(self):
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        return f"""# --- Categorical Analysis ---

# Frequency table
freq_table = pd.crosstab(
    analysis_data[CONFIG['trt_var']],
    analysis_data['AVALC'],
    margins=True
)
print("\\nFrequency Table:")
print(freq_table)

# Chi-square test
contingency = pd.crosstab(analysis_data[CONFIG['trt_var']], analysis_data['AVALC'])
chi2, p_value, dof, expected = chi2_contingency(contingency)
print(f"\\nChi-square test: chi2={{chi2:.4f}}, p={{p_value:.4f}}, df={{dof}}")

# Fisher's exact test (for 2x2 tables)
if contingency.shape == (2, 2):
    odds_ratio, fisher_p = fisher_exact(contingency)
    print(f"Fisher's exact test: OR={{odds_ratio:.4f}}, p={{fisher_p:.4f}}")"""

    def _gen_pk_py(self):
        return """# --- Pharmacokinetic Analysis ---

pk_summary = analysis_data.groupby(
    [CONFIG['trt_var'], 'PARAM', 'PARAMCD']
)['AVAL'].agg([
    ('n', 'count'),
    ('mean', 'mean'),
    ('std', 'std'),
    ('median', 'median'),
    ('min', 'min'),
    ('max', 'max')
]).reset_index()

# Geometric mean and CV%
pk_summary['gmean'] = analysis_data.groupby(
    [CONFIG['trt_var'], 'PARAM', 'PARAMCD']
)['AVAL'].apply(lambda x: np.exp(np.log(x[x > 0]).mean())).values

pk_summary['cv_pct'] = (pk_summary['std'] / pk_summary['mean']) * 100

print("\\nPK Parameter Summary:")
print(pk_summary)"""

    def _gen_subgroup_py(self):
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        subgroups = self.analysis_params.get('subgroup_variables', ['SEX', 'AGEGR1', 'RACE'])

        return f"""# --- Subgroup Analysis ---

subgroup_results = []

for sg_var in {subgroups}:
    for category in analysis_data[sg_var].unique():
        subset = analysis_data[analysis_data[sg_var] == category]
        
        if len(subset[CONFIG['trt_var']].unique()) < 2:
            continue
        
        formula = f'CHG ~ C({{CONFIG["trt_var"]}}) + BASE'
        try:
            model = ols(formula, data=subset).fit()
            
            # Get treatment effect
            params = model.params
            conf_int = model.conf_int()
            
            for param_name in params.index:
                if CONFIG['trt_var'] in param_name:
                    subgroup_results.append({{
                        'subgroup': sg_var,
                        'category': category,
                        'estimate': params[param_name],
                        'lower_ci': conf_int.loc[param_name, 0],
                        'upper_ci': conf_int.loc[param_name, 1],
                        'p_value': model.pvalues[param_name]
                    }})
        except Exception as e:
            print(f"Warning: Could not fit model for {{sg_var}}={{category}}: {{e}}")

sg_df = pd.DataFrame(subgroup_results)
print("\\nSubgroup Analysis Results:")
print(sg_df)

# Forest plot
if not sg_df.empty:
    fig, ax = plt.subplots(figsize=(10, len(sg_df) * 0.5 + 2))
    
    y_pos = range(len(sg_df))
    ax.errorbarh(
        y_pos, sg_df['estimate'],
        xerr=[sg_df['estimate'] - sg_df['lower_ci'],
              sg_df['upper_ci'] - sg_df['estimate']],
        fmt='o', color='navy', capsize=3
    )
    ax.axvline(x=0, color='grey', linestyle='--', alpha=0.7)
    ax.set_yticks(y_pos)
    ax.set_yticklabels([f"{{r['subgroup']}}: {{r['category']}}" for _, r in sg_df.iterrows()])
    ax.set_xlabel('Treatment Difference (95% CI)')
    ax.set_title('Subgroup Analysis - Forest Plot')
    plt.tight_layout()"""

    def _generate_output(self):
        output_generators = {
            'table': self._gen_table_output_py,
            'figure': self._gen_figure_output_py,
            'listing': self._gen_listing_output_py,
            'dataset': self._gen_dataset_output_py
        }
        generator = output_generators.get(self.output_type, self._gen_table_output_py)
        return generator()

    def _gen_table_output_py(self):
        return f"""# --- Generate Table Output ---

output_file = CONFIG['output_path'] / f"{{CONFIG['output_name']}}.{self.output_data.get('format', 'xlsx')}"

# Create formatted output table
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Results"

# Header information
ws['A1'] = CONFIG['study_id']
ws['A1'].font = Font(bold=True, size=12)
ws['A2'] = f"Protocol: {{CONFIG['protocol']}}"
ws['A3'] = f"Table X.X.X: {self._get_purpose_description()}"
ws['A3'].font = Font(bold=True, size=11)
ws['A4'] = f"Population: {self.analysis_params.get('population', 'Safety').title()} Population"

# Column headers
headers = ['Statistic', 'Placebo\\n(N={{}})', 'Treatment A\\n(N={{}})', 'Treatment B\\n(N={{}})']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Write results data
# (Implementation would write actual results from analysis)

# Footnotes
last_row = ws.max_row + 2
ws.cell(row=last_row, column=1, value=f"Source: {', '.join(self.input_data.get('datasets', ['adsl']))}")
ws.cell(row=last_row + 1, column=1, value=f"Program: {{CONFIG['prog_name']}}.py")
ws.cell(row=last_row + 2, column=1, value=f"Generated: {{datetime.now().strftime('%d%b%Y %H:%M')}}")

wb.save(str(output_file))
print(f"NOTE: Table written to {{output_file}}")"""

    def _gen_figure_output_py(self):
        return f"""# --- Generate Figure Output ---

output_file = CONFIG['output_path'] / f"{{CONFIG['output_name']}}.pdf"

# Finalize figure
plt.suptitle('Figure X.X.X', fontsize=12, fontweight='bold', y=0.98)
plt.title('{self._get_purpose_description()}', fontsize=10, pad=20)

# Add footnotes
plt.figtext(0.1, 0.02, 
    f"Source: {', '.join(self.input_data.get('datasets', ['adsl']))}\\n"
    f"Program: {{CONFIG['prog_name']}}.py\\n"
    f"Generated: {{datetime.now().strftime('%d%b%Y')}}",
    fontsize=7, va='bottom')

plt.tight_layout(rect=[0, 0.05, 1, 0.95])
plt.savefig(str(output_file), dpi=300, bbox_inches='tight')
plt.close()

print(f"NOTE: Figure written to {{output_file}}")"""

    def _gen_listing_output_py(self):
        return f"""# --- Generate Listing Output ---

output_file = CONFIG['output_path'] / f"{{CONFIG['output_name']}}.xlsx"

listing_cols = ['USUBJID', 'SITEID'] + {self.variables.get('analysis_variables', ['AVAL'])}
listing_data = analysis_data[listing_cols].sort_values('USUBJID')

listing_data.to_excel(str(output_file), index=False, sheet_name='Listing')
print(f"NOTE: Listing written to {{output_file}}")"""

    def _gen_dataset_output_py(self):
        return f"""# --- Generate Analysis Dataset ---

output_file = CONFIG['output_path'] / f"{{CONFIG['output_name']}}.xpt"

# Export as XPT (SAS Transport) for submission
pyreadstat.write_xport(
    final_dataset,
    str(output_file),
    file_label='{self.submission.get('study_id', 'STUDY001')} Analysis Dataset'
)

print(f"NOTE: Dataset written to {{output_file}}")
print(f"NOTE: Dataset has {{final_dataset.shape[0]}} records and {{final_dataset.shape[1]}} variables.")"""

    def _generate_validation(self):
        return f"""# --- Validation Checks ---

print("\\n" + "=" * 60)
print("VALIDATION RESULTS")
print("=" * 60)

# Check for missing critical variables
n_missing_usubjid = analysis_data['USUBJID'].isna().sum()
n_missing_trt = analysis_data[CONFIG['trt_var']].isna().sum()
n_subjects = analysis_data['USUBJID'].nunique()

print(f"  Missing USUBJID: {{n_missing_usubjid}}")
print(f"  Missing Treatment: {{n_missing_trt}}")
print(f"  Total unique subjects: {{n_subjects}}")
print(f"  Total records: {{analysis_data.shape[0]}}")

# Verify population counts
assert n_missing_usubjid == 0, "ERROR: Missing USUBJID values found!"
assert n_missing_trt == 0, "ERROR: Missing treatment values found!"

print("\\n" + "=" * 60)
print(f"Program: {{CONFIG['prog_name']}}.py")
print(f"Output:  {{CONFIG['output_name']}}.{self.output_data.get('format', 'rtf')}")
print(f"Status:  COMPLETE")
print(f"Time:    {{datetime.now().strftime('%d%b%Y %H:%M:%S')}}")
print("=" * 60)"""

    # Helper methods
    def _get_population_flag(self):
        pop_flags = {
            'safety': 'SAFFL', 'itt': 'ITTFL', 'per_protocol': 'PPROTFL',
            'pk': 'PKFL', 'full_analysis': 'FASFL'
        }
        return pop_flags.get(self.analysis_params.get('population', 'safety'), 'SAFFL')

    def _get_purpose_description(self):
        descriptions = {
            'descriptive': 'Generate descriptive statistics summary',
            'efficacy': 'Primary efficacy endpoint analysis',
            'safety': 'Safety summary of adverse events',
            'survival': 'Time-to-event survival analysis',
            'mixed_model': 'Mixed model repeated measures analysis',
            'categorical': 'Categorical data analysis',
            'pk': 'Pharmacokinetic parameter summary',
            'subgroup': 'Subgroup analysis with forest plot'
        }
        return descriptions.get(self.analysis_type, 'Statistical analysis')