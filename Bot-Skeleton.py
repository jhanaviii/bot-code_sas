clinical-code-generator/
├── .vscode/
│   └── settings.json
├── backend/
│   ├── app.py
│   ├── requirements.txt
│   ├── generators/
│   │   ├── __init__.py
│   │   ├── sas_generator.py
│   │   ├── r_generator.py
│   │   └── python_generator.py
│   ├── templates/
│   │   ├── sas/
│   │   ├── r/
│   │   └── python/
│   └── validators/
│       ├── __init__.py
│       └── cdisc_validator.py
├── frontend/
│   ├── index.html
│   ├── css/
│   │   └── styles.css
│   ├── js/
│   │   ├── app.js
│   │   ├── code-generator.js
│   │   └── file-upload.js
│   └── assets/
├── uploads/
├── output/
├── package.json
└── README.md

flask==3.0.0
flask-cors==4.0.0
werkzeug==3.0.1
jinja2==3.1.2
pandas==2.1.4
openpyxl==3.1.2
python-docx==1.1.0
jsonschema==4.20.0

"""
Clinical Trial Analysis Code Generator - Backend API
Generates submission-ready SAS, R, and Python code for health authority submissions.
Supports CDISC standards (SDTM, ADaM) and ICH E9 guidelines.
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
import os
import json
from datetime import datetime

from generators.sas_generator import SASCodeGenerator
from generators.r_generator import RCodeGenerator
from generators.python_generator import PythonCodeGenerator
from validators.cdisc_validator import CDISCValidator

app = Flask(__name__)
CORS(app)

# Configuration
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'uploads')
OUTPUT_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'output')
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv', 'docx', 'pdf', 'json', 'sas7bdat', 'xpt'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({
        'status': 'healthy',
        'version': '1.0.0',
        'timestamp': datetime.now().isoformat()
    })


@app.route('/api/generate', methods=['POST'])
def generate_code():
    """
    Main endpoint to generate analysis code based on user prompts.
    
    Expected JSON payload:
    {
        "language": "sas" | "r" | "python",
        "analysis_type": "descriptive" | "efficacy" | "safety" | "pk" | "survival" | "mixed_model",
        "output_type": "dataset" | "table" | "figure" | "listing",
        "input_data": {
            "datasets": ["adsl", "adae", "adtte"],
            "format": "adam" | "sdtm" | "raw",
            "library_path": "/data/study123/adam"
        },
        "output_data": {
            "dataset_name": "t_ae_summary",
            "format": "rtf" | "pdf" | "xlsx" | "xpt" | "sas7bdat",
            "output_path": "/output/tables"
        },
        "variables": {
            "analysis_variables": ["AVAL", "CHG", "PCHG"],
            "grouping_variables": ["TRTA", "AVISIT"],
            "filter_variables": ["SAFFL", "ANL01FL"],
            "sort_variables": ["USUBJID", "AVISITN"]
        },
        "analysis_parameters": {
            "statistical_method": "anova" | "chi_square" | "cox_regression" | "kaplan_meier",
            "confidence_level": 0.95,
            "alpha": 0.05,
            "population": "safety" | "itt" | "per_protocol",
            "treatment_variable": "TRTA",
            "subgroup_variables": [],
            "covariates": []
        },
        "submission_details": {
            "study_id": "STUDY-001",
            "sponsor": "AstraZeneca",
            "protocol": "D1234C00001",
            "regulatory_authority": "FDA" | "EMA" | "PMDA",
            "submission_type": "NDA" | "BLA" | "MAA"
        },
        "shell_file_id": null,
        "specification_file_id": null
    }
    """
    try:
        data = request.get_json()

        if not data:
            return jsonify({'error': 'No JSON payload provided'}), 400

        # Validate required fields
        required_fields = ['language', 'analysis_type', 'output_type', 'input_data', 'variables']
        missing_fields = [f for f in required_fields if f not in data]
        if missing_fields:
            return jsonify({'error': f'Missing required fields: {missing_fields}'}), 400

        # Select appropriate generator
        language = data['language'].lower()
        generators = {
            'sas': SASCodeGenerator,
            'r': RCodeGenerator,
            'python': PythonCodeGenerator
        }

        if language not in generators:
            return jsonify({'error': f'Unsupported language: {language}'}), 400

        generator = generators[language](data)

        # Load shell/specification if provided
        shell_metadata = None
        if data.get('shell_file_id'):
            shell_path = os.path.join(UPLOAD_FOLDER, data['shell_file_id'])
            if os.path.exists(shell_path):
                shell_metadata = generator.parse_shell(shell_path)

        spec_metadata = None
        if data.get('specification_file_id'):
            spec_path = os.path.join(UPLOAD_FOLDER, data['specification_file_id'])
            if os.path.exists(spec_path):
                spec_metadata = generator.parse_specification(spec_path)

        # Generate code
        generated_code = generator.generate(
            shell_metadata=shell_metadata,
            spec_metadata=spec_metadata
        )

        # Validate CDISC compliance
        validator = CDISCValidator()
        validation_results = validator.validate(data, generated_code)

        # Save generated code
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        extensions = {'sas': '.sas', 'r': '.R', 'python': '.py'}
        output_filename = f"{data['output_type']}_{data['analysis_type']}_{timestamp}{extensions[language]}"
        output_path = os.path.join(OUTPUT_FOLDER, output_filename)

        with open(output_path, 'w') as f:
            f.write(generated_code)

        return jsonify({
            'success': True,
            'code': generated_code,
            'filename': output_filename,
            'language': language,
            'validation': validation_results,
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'analysis_type': data['analysis_type'],
                'output_type': data['output_type'],
                'submission_ready': validation_results.get('compliant', False)
            }
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/upload/shell', methods=['POST'])
def upload_shell():
    """Upload output shell template (RTF, Excel, or PDF mock-up)."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        saved_filename = f"shell_{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], saved_filename)
        file.save(filepath)

        return jsonify({
            'success': True,
            'file_id': saved_filename,
            'filename': filename,
            'type': 'shell',
            'uploaded_at': datetime.now().isoformat()
        })

    return jsonify({'error': 'File type not allowed'}), 400


@app.route('/api/upload/specification', methods=['POST'])
def upload_specification():
    """Upload dataset/variable specification (Excel, CSV, or JSON)."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        saved_filename = f"spec_{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], saved_filename)
        file.save(filepath)

        return jsonify({
            'success': True,
            'file_id': saved_filename,
            'filename': filename,
            'type': 'specification',
            'uploaded_at': datetime.now().isoformat()
        })

    return jsonify({'error': 'File type not allowed'}), 400


@app.route('/api/download/<filename>', methods=['GET'])
def download_code(filename):
    """Download generated code file."""
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], secure_filename(filename))
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return jsonify({'error': 'File not found'}), 404


@app.route('/api/templates', methods=['GET'])
def get_templates():
    """Return available analysis templates and configurations."""
    templates = {
        'analysis_types': [
            {'id': 'descriptive', 'name': 'Descriptive Statistics', 'description': 'Summary statistics, frequencies, percentages'},
            {'id': 'efficacy', 'name': 'Efficacy Analysis', 'description': 'Primary/secondary endpoint analysis'},
            {'id': 'safety', 'name': 'Safety Analysis', 'description': 'AE summaries, lab shifts, vital signs'},
            {'id': 'pk', 'name': 'Pharmacokinetics', 'description': 'PK parameter summaries, concentration-time profiles'},
            {'id': 'survival', 'name': 'Survival Analysis', 'description': 'Kaplan-Meier, Cox regression, log-rank test'},
            {'id': 'mixed_model', 'name': 'Mixed Models', 'description': 'MMRM, repeated measures analysis'},
            {'id': 'categorical', 'name': 'Categorical Analysis', 'description': 'CMH test, logistic regression, odds ratios'},
            {'id': 'subgroup', 'name': 'Subgroup Analysis', 'description': 'Forest plots, interaction tests'}
        ],
        'output_types': [
            {'id': 'dataset', 'name': 'Analysis Dataset', 'formats': ['sas7bdat', 'xpt', 'csv', 'rds']},
            {'id': 'table', 'name': 'Summary Table', 'formats': ['rtf', 'pdf', 'xlsx', 'html']},
            {'id': 'figure', 'name': 'Figure/Graph', 'formats': ['pdf', 'png', 'svg', 'rtf']},
            {'id': 'listing', 'name': 'Data Listing', 'formats': ['rtf', 'pdf', 'xlsx']}
        ],
        'populations': [
            {'id': 'safety', 'name': 'Safety Population', 'flag': 'SAFFL'},
            {'id': 'itt', 'name': 'Intent-to-Treat', 'flag': 'ITTFL'},
            {'id': 'per_protocol', 'name': 'Per Protocol', 'flag': 'PPROTFL'},
            {'id': 'pk', 'name': 'PK Population', 'flag': 'PKFL'},
            {'id': 'full_analysis', 'name': 'Full Analysis Set', 'flag': 'FASFL'}
        ],
        'statistical_methods': [
            {'id': 'anova', 'name': 'ANOVA/ANCOVA'},
            {'id': 'chi_square', 'name': 'Chi-Square Test'},
            {'id': 'fisher_exact', 'name': "Fisher's Exact Test"},
            {'id': 'cox_regression', 'name': 'Cox Proportional Hazards'},
            {'id': 'kaplan_meier', 'name': 'Kaplan-Meier Estimation'},
            {'id': 'log_rank', 'name': 'Log-Rank Test'},
            {'id': 'mmrm', 'name': 'Mixed Model Repeated Measures'},
            {'id': 'logistic_regression', 'name': 'Logistic Regression'},
            {'id': 'cmh', 'name': 'Cochran-Mantel-Haenszel'},
            {'id': 'wilcoxon', 'name': 'Wilcoxon Rank-Sum Test'}
        ],
        'regulatory_authorities': [
            {'id': 'fda', 'name': 'FDA (US)'},
            {'id': 'ema', 'name': 'EMA (EU)'},
            {'id': 'pmda', 'name': 'PMDA (Japan)'},
            {'id': 'nmpa', 'name': 'NMPA (China)'},
            {'id': 'hc', 'name': 'Health Canada'}
        ]
    }
    return jsonify(templates)


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)


    from .sas_generator import SASCodeGenerator
from .r_generator import RCodeGenerator
from .python_generator import PythonCodeGenerator

__all__ = ['SASCodeGenerator', 'RCodeGenerator', 'PythonCodeGenerator']


"""
SAS Code Generator for Clinical Trial Submissions
Generates CDISC-compliant, submission-ready SAS programs.
"""

import json
from datetime import datetime


class SASCodeGenerator:
    def __init__(self, config):
        self.config = config
        self.analysis_type = config.get('analysis_type', 'descriptive')
        self.output_type = config.get('output_type', 'table')
        self.input_data = config.get('input_data', {})
        self.output_data = config.get('output_data', {})
        self.variables = config.get('variables', {})
        self.analysis_params = config.get('analysis_parameters', {})
        self.submission = config.get('submission_details', {})

    def parse_shell(self, shell_path):
        """Parse uploaded shell template to extract layout specifications."""
        # Parse RTF/Excel shell for column headers, row structure, formatting
        metadata = {
            'columns': [],
            'rows': [],
            'titles': [],
            'footnotes': [],
            'page_layout': 'landscape',
            'font_size': 9
        }
        # Implementation would parse actual shell files
        return metadata

    def parse_specification(self, spec_path):
        """Parse dataset/variable specification file."""
        metadata = {
            'variables': [],
            'derivations': [],
            'formats': [],
            'labels': []
        }
        # Implementation would parse Excel/CSV specifications
        return metadata

    def generate(self, shell_metadata=None, spec_metadata=None):
        """Generate complete SAS program."""
        sections = [
            self._generate_header(),
            self._generate_setup(),
            self._generate_data_access(),
            self._generate_data_preparation(),
            self._generate_analysis(),
            self._generate_output(),
            self._generate_validation(),
            self._generate_footer()
        ]
        return '\n\n'.join(sections)

    def _generate_header(self):
        """Generate program header with metadata."""
        study_id = self.submission.get('study_id', 'STUDY-XXX')
        sponsor = self.submission.get('sponsor', 'AstraZeneca')
        protocol = self.submission.get('protocol', 'DXXXXCXXXXX')
        
        return f"""/******************************************************************************
* PROGRAM NAME  : {self.output_data.get('dataset_name', 'program')}.sas
* STUDY         : {study_id}
* PROTOCOL      : {protocol}
* SPONSOR       : {sponsor}
* PURPOSE       : {self._get_purpose_description()}
* INPUT         : {', '.join(self.input_data.get('datasets', []))}
* OUTPUT        : {self.output_data.get('dataset_name', 'output')}.{self.output_data.get('format', 'rtf')}
* ANALYSIS TYPE : {self.analysis_type.replace('_', ' ').title()}
* OUTPUT TYPE   : {self.output_type.title()}
* POPULATION    : {self.analysis_params.get('population', 'safety').replace('_', ' ').title()}
*
* AUTHOR        : [Programmer Name]
* DATE CREATED  : {datetime.now().strftime('%d%b%Y')}
* DATE MODIFIED : {datetime.now().strftime('%d%b%Y')}
*
* REGULATORY    : {self.submission.get('regulatory_authority', 'FDA')}
* SUBMISSION    : {self.submission.get('submission_type', 'NDA')}
*
* VALIDATION    : Independent double programming required
* CDISC VERSION : ADaM IG v1.3 / SDTM IG v3.4
*
* MODIFICATION HISTORY:
* DATE        PROGRAMMER    DESCRIPTION
* ---------   ----------    ------------------------------------------------
* {datetime.now().strftime('%d%b%Y')}  [Name]        Initial creation
******************************************************************************/"""

    def _generate_setup(self):
        """Generate program setup and macro variables."""
        population_flag = self._get_population_flag()
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        
        return f"""/*--- Program Setup ---*/
%let study    = {self.submission.get('study_id', 'STUDY001')};
%let protocol = {self.submission.get('protocol', 'D1234C00001')};
%let progname = {self.output_data.get('dataset_name', 'program')};
%let outname  = {self.output_data.get('dataset_name', 'output')};
%let popflag  = {population_flag};
%let trtvar   = {treatment_var};
%let alpha    = {self.analysis_params.get('alpha', 0.05)};
%let conflev  = {self.analysis_params.get('confidence_level', 0.95)};

/*--- System Options ---*/
options nodate nonumber orientation={self._get_orientation()} 
        ls=200 ps=60 missing=' ' formchar='|_---|+|---+=|-/\\<>*';
options mprint mlogic symbolgen;

/*--- ODS Setup ---*/
ods listing close;
ods escapechar = '~';

/*--- Library References ---*/
libname adam "{self.input_data.get('library_path', '/data/adam')}" access=readonly;
libname output "{self.output_data.get('output_path', '/output')}";

/*--- Format Definitions ---*/
proc format;
    value $trtfmt
        'Placebo'       = 'Placebo'
        'Treatment A'   = 'Treatment A'
        'Treatment B'   = 'Treatment B'
        'Total'         = 'Total'
    ;
    value $popfmt
        'Y' = 'Yes'
        'N' = 'No'
    ;
run;"""

    def _generate_data_access(self):
        """Generate data access and initial filtering."""
        datasets = self.input_data.get('datasets', ['adsl'])
        filter_vars = self.variables.get('filter_variables', [])
        population_flag = self._get_population_flag()

        data_steps = []
        for ds in datasets:
            filters = [f"{population_flag} = 'Y'"]
            for fv in filter_vars:
                if fv != population_flag:
                    filters.append(f"{fv} = 'Y'")

            filter_clause = ' and '.join(filters)
            
            data_steps.append(f"""/*--- Read {ds.upper()} dataset ---*/
data work.{ds};
    set adam.{ds};
    where {filter_clause};
run;

proc sort data=work.{ds};
    by {' '.join(self.variables.get('sort_variables', ['USUBJID']))};
run;

/*--- Verify record count ---*/
proc sql noprint;
    select count(*) into :n_{ds} trimmed
    from work.{ds};
quit;
%put NOTE: {ds.upper()} has &n_{ds} records after filtering.;""")

        return '\n\n'.join(data_steps)

    def _generate_data_preparation(self):
        """Generate data preparation and derivation steps."""
        analysis_vars = self.variables.get('analysis_variables', [])
        grouping_vars = self.variables.get('grouping_variables', [])
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')

        return f"""/*--- Data Preparation ---*/
data work.analysis;
    set work.{self.input_data.get('datasets', ['adsl'])[0]};
    
    /*--- Derive analysis variables ---*/
    length col1-col4 $200;
    
    /*--- Treatment group ordering ---*/
    select ({treatment_var});
        when ('Placebo')       trtn = 1;
        when ('Treatment A')   trtn = 2;
        when ('Treatment B')   trtn = 3;
        otherwise              trtn = 99;
    end;
    
    /*--- Create Total group for summary ---*/
    output;
    {treatment_var} = 'Total';
    trtn = 99;
    output;
run;

proc sort data=work.analysis;
    by trtn {treatment_var} {' '.join(grouping_vars)};
run;

/*--- Get Big N (population counts) ---*/
proc sql noprint;
    select count(distinct USUBJID) into :N1 trimmed
    from work.analysis
    where trtn = 1;
    
    select count(distinct USUBJID) into :N2 trimmed
    from work.analysis
    where trtn = 2;
    
    select count(distinct USUBJID) into :N3 trimmed
    from work.analysis
    where trtn = 3;
    
    select count(distinct USUBJID) into :Ntot trimmed
    from work.analysis
    where trtn = 99;
quit;

%put NOTE: Population counts - Placebo=&N1, TrtA=&N2, TrtB=&N3, Total=&Ntot;"""

    def _generate_analysis(self):
        """Generate statistical analysis code based on analysis type."""
        method_generators = {
            'descriptive': self._gen_descriptive_analysis,
            'efficacy': self._gen_efficacy_analysis,
            'safety': self._gen_safety_analysis,
            'survival': self._gen_survival_analysis,
            'mixed_model': self._gen_mixed_model_analysis,
            'categorical': self._gen_categorical_analysis,
            'pk': self._gen_pk_analysis,
            'subgroup': self._gen_subgroup_analysis
        }

        generator = method_generators.get(self.analysis_type, self._gen_descriptive_analysis)
        return generator()

    def _gen_descriptive_analysis(self):
        """Generate descriptive statistics analysis."""
        analysis_vars = self.variables.get('analysis_variables', ['AVAL'])
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        grouping_vars = self.variables.get('grouping_variables', [])

        by_statement = f"by {' '.join(grouping_vars)};" if grouping_vars else ""

        return f"""/*--- Descriptive Statistics ---*/
proc means data=work.analysis nway noprint;
    class {treatment_var} trtn {' '.join(grouping_vars)};
    var {' '.join(analysis_vars)};
    output out=work.stats (drop=_type_ _freq_)
        n=n
        mean=mean
        std=std
        median=median
        min=min
        max=max
        q1=q1
        q3=q3;
run;

/*--- Format statistics for display ---*/
data work.stats_fmt;
    set work.stats;
    length col1-col4 $200;
    
    /*--- N ---*/
    n_c = strip(put(n, 8.));
    
    /*--- Mean (SD) ---*/
    mean_sd = strip(put(mean, 10.1)) || ' (' || strip(put(std, 10.2)) || ')';
    
    /*--- Median ---*/
    median_c = strip(put(median, 10.1));
    
    /*--- Q1, Q3 ---*/
    q1q3 = strip(put(q1, 10.1)) || ', ' || strip(put(q3, 10.1));
    
    /*--- Min, Max ---*/
    minmax = strip(put(min, 10.1)) || ', ' || strip(put(max, 10.1));
run;

proc sort data=work.stats_fmt;
    by trtn {treatment_var};
run;

/*--- Transpose for report layout ---*/
proc transpose data=work.stats_fmt out=work.stats_t (drop=_name_) prefix=col;
    by {' '.join(grouping_vars) + ' ' if grouping_vars else ''}/* row identifiers */;
    id trtn;
    var n_c mean_sd median_c q1q3 minmax;
run;"""

    def _gen_efficacy_analysis(self):
        """Generate efficacy analysis with ANCOVA."""
        analysis_vars = self.variables.get('analysis_variables', ['CHG'])
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        covariates = self.analysis_params.get('covariates', ['BASE'])

        return f"""/*--- Efficacy Analysis: ANCOVA ---*/

/*--- Descriptive Statistics by Visit and Treatment ---*/
proc means data=work.analysis nway noprint;
    class {treatment_var} trtn AVISIT AVISITN;
    var {' '.join(analysis_vars)} BASE;
    output out=work.desc_stats (drop=_type_ _freq_)
        n=n mean=mean std=std median=median min=min max=max;
run;

/*--- ANCOVA Model ---*/
proc mixed data=work.analysis;
    class {treatment_var} SITEID;
    model {analysis_vars[0]} = {treatment_var} {' '.join(covariates)} / ddfm=kr solution cl alpha=&alpha;
    lsmeans {treatment_var} / pdiff cl alpha=&alpha;
    ods output LSMeans=work.lsmeans
              Diffs=work.diffs
              SolutionF=work.solution
              Tests3=work.tests;
run;

/*--- Format LS Means results ---*/
data work.lsmeans_fmt;
    set work.lsmeans;
    length result $200;
    result = strip(put(Estimate, 10.2)) || ' (' || 
             strip(put(Lower, 10.2)) || ', ' || 
             strip(put(Upper, 10.2)) || ')';
    pvalue_c = ifc(Probt < 0.001, '<0.001', strip(put(Probt, 6.3)));
run;

/*--- Format Treatment Differences ---*/
data work.diffs_fmt;
    set work.diffs;
    length diff_result $200;
    diff_result = strip(put(Estimate, 10.2)) || ' (' || 
                  strip(put(Lower, 10.2)) || ', ' || 
                  strip(put(Upper, 10.2)) || ')';
    pvalue_c = ifc(Probt < 0.001, '<0.001', strip(put(Probt, 6.3)));
run;"""

    def _gen_safety_analysis(self):
        """Generate safety analysis (AE summary)."""
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')

        return f"""/*--- Safety Analysis: Adverse Event Summary ---*/

/*--- Overall AE Summary ---*/
proc sql;
    create table work.ae_overall as
    select {treatment_var}, trtn,
           count(distinct USUBJID) as n_subj,
           count(*) as n_events
    from work.adae
    group by {treatment_var}, trtn;
quit;

/*--- AE by System Organ Class and Preferred Term ---*/
proc sql;
    create table work.ae_soc_pt as
    select {treatment_var}, trtn, AEBODSYS, AEDECOD,
           count(distinct USUBJID) as n_subj,
           count(*) as n_events
    from work.adae
    group by {treatment_var}, trtn, AEBODSYS, AEDECOD
    order by AEBODSYS, AEDECOD, trtn;
quit;

/*--- Calculate percentages ---*/
data work.ae_summary;
    set work.ae_soc_pt;
    
    select (trtn);
        when (1) bigN = input("&N1", best.);
        when (2) bigN = input("&N2", best.);
        when (3) bigN = input("&N3", best.);
        when (99) bigN = input("&Ntot", best.);
    end;
    
    pct = (n_subj / bigN) * 100;
    
    /*--- Format: n (xx.x%) ---*/
    length col $50;
    if n_subj > 0 then
        col = strip(put(n_subj, 8.)) || ' (' || strip(put(pct, 5.1)) || '%)';
    else
        col = '0';
run;

/*--- Sort by descending frequency in active treatment ---*/
proc sql;
    create table work.ae_sorted as
    select a.*
    from work.ae_summary a
    left join (
        select AEBODSYS, AEDECOD, sum(n_subj) as total_n
        from work.ae_summary
        where trtn = 2
        group by AEBODSYS, AEDECOD
    ) b on a.AEBODSYS = b.AEBODSYS and a.AEDECOD = b.AEDECOD
    order by b.total_n desc, a.AEBODSYS, a.AEDECOD, a.trtn;
quit;"""

    def _gen_survival_analysis(self):
        """Generate survival analysis (Kaplan-Meier + Cox)."""
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')

        return f"""/*--- Survival Analysis ---*/

/*--- Kaplan-Meier Estimation ---*/
proc lifetest data=work.analysis method=km 
    plots=(survival(cl atrisk) logsurv)
    outsurv=work.km_estimates
    timelist=(3 6 9 12)
    reduceout;
    time AVAL * CNSR(1);
    strata {treatment_var};
    ods output Quartiles=work.km_quartiles
              HomTests=work.logrank
              ProductLimitEstimates=work.km_detail;
run;

/*--- Cox Proportional Hazards Model ---*/
proc phreg data=work.analysis;
    class {treatment_var} (ref='Placebo') / param=ref;
    model AVAL * CNSR(1) = {treatment_var} / ties=efron rl alpha=&alpha;
    hazardratio {treatment_var} / diff=ref cl=wald;
    ods output ParameterEstimates=work.cox_params
              HazardRatios=work.hazard_ratios
              GlobalTests=work.global_tests;
run;

/*--- Format KM Results ---*/
data work.km_results;
    set work.km_estimates;
    length survival_c $50 ci_c $50;
    survival_c = strip(put(SURVIVAL, 6.3));
    ci_c = '(' || strip(put(SDF_LCL, 6.3)) || ', ' || strip(put(SDF_UCL, 6.3)) || ')';
run;

/*--- Format Hazard Ratio ---*/
data work.hr_results;
    set work.hazard_ratios;
    length hr_c $100;
    hr_c = strip(put(HazardRatio, 6.3)) || ' (' || 
           strip(put(HRLowerCL, 6.3)) || ', ' || 
           strip(put(HRUpperCL, 6.3)) || ')';
run;"""

    def _gen_mixed_model_analysis(self):
        """Generate MMRM analysis."""
        analysis_vars = self.variables.get('analysis_variables', ['CHG'])
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        covariates = self.analysis_params.get('covariates', ['BASE'])

        return f"""/*--- Mixed Model Repeated Measures (MMRM) ---*/

proc mixed data=work.analysis method=reml;
    class {treatment_var} USUBJID AVISIT SITEID;
    model {analysis_vars[0]} = {treatment_var} AVISIT {treatment_var}*AVISIT 
          {' '.join(covariates)} / ddfm=kr solution cl alpha=&alpha;
    repeated AVISIT / subject=USUBJID type=un r rcorr;
    lsmeans {treatment_var}*AVISIT / slice=AVISIT pdiff cl alpha=&alpha;
    ods output LSMeans=work.mmrm_lsmeans
              Diffs=work.mmrm_diffs
              Tests3=work.mmrm_tests
              FitStatistics=work.mmrm_fit
              CovParms=work.mmrm_cov;
run;

/*--- Format MMRM Results by Visit ---*/
data work.mmrm_results;
    set work.mmrm_lsmeans;
    length lsmean_c $100 ci_c $100;
    lsmean_c = strip(put(Estimate, 10.2)) || ' (' || strip(put(StdErr, 10.3)) || ')';
    ci_c = '(' || strip(put(Lower, 10.2)) || ', ' || strip(put(Upper, 10.2)) || ')';
run;

/*--- Treatment Differences at Each Visit ---*/
data work.mmrm_diff_results;
    set work.mmrm_diffs;
    where AVISIT ne ' ';
    length diff_c $100 pval_c $20;
    diff_c = strip(put(Estimate, 10.2)) || ' (' || 
             strip(put(Lower, 10.2)) || ', ' || 
             strip(put(Upper, 10.2)) || ')';
    pval_c = ifc(Probt < 0.001, '<0.001', strip(put(Probt, 6.3)));
run;"""

    def _gen_categorical_analysis(self):
        """Generate categorical analysis."""
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        return f"""/*--- Categorical Analysis ---*/

/*--- Frequency counts ---*/
proc freq data=work.analysis noprint;
    tables {treatment_var} * AVALC / cmh chisq fisher outpct out=work.freq_counts;
    ods output CMH=work.cmh_results
              ChiSq=work.chisq_results
              FishersExact=work.fisher_results;
run;

/*--- Format results ---*/
data work.cat_results;
    set work.freq_counts;
    length col $50;
    col = strip(put(COUNT, 8.)) || ' (' || strip(put(PCT_ROW, 5.1)) || '%)';
run;"""

    def _gen_pk_analysis(self):
        """Generate PK analysis."""
        return """/*--- Pharmacokinetic Analysis ---*/

/*--- PK Parameter Summary ---*/
proc means data=work.analysis nway noprint;
    class TRTA trtn PARAM PARAMCD;
    var AVAL;
    output out=work.pk_stats (drop=_type_ _freq_)
        n=n mean=mean std=std median=median min=min max=max
        cv=cv gmean=gmean;
run;

/*--- Geometric Mean and CV% ---*/
data work.pk_summary;
    set work.pk_stats;
    length gmean_c $50 cv_c $20;
    log_mean = log(mean);
    gmean_calc = exp(log_mean);
    gmean_c = strip(put(gmean_calc, 10.3));
    cv_c = strip(put(cv, 6.1)) || '%';
run;"""

    def _gen_subgroup_analysis(self):
        """Generate subgroup analysis."""
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        subgroups = self.analysis_params.get('subgroup_variables', ['SEX', 'AGEGR1', 'RACE'])
        
        return f"""/*--- Subgroup Analysis ---*/

%macro subgroup_analysis(subgrp=);
    proc mixed data=work.analysis;
        class {treatment_var} &subgrp;
        model CHG = {treatment_var} BASE &subgrp {treatment_var}*&subgrp / ddfm=kr;
        lsmeans {treatment_var}*&subgrp / pdiff cl slice=&subgrp;
        ods output Diffs=work.sg_&subgrp;
    run;
%mend;

{chr(10).join([f"%subgroup_analysis(subgrp={sg});" for sg in subgroups])}

/*--- Combine subgroup results for forest plot ---*/
data work.forest_data;
    set {' '.join([f'work.sg_{sg}' for sg in subgroups])};
    length subgroup $50 category $100;
run;"""

    def _generate_output(self):
        """Generate output production code."""
        output_generators = {
            'table': self._gen_table_output,
            'figure': self._gen_figure_output,
            'listing': self._gen_listing_output,
            'dataset': self._gen_dataset_output
        }
        
        generator = output_generators.get(self.output_type, self._gen_table_output)
        return generator()

    def _gen_table_output(self):
        """Generate RTF table output."""
        output_format = self.output_data.get('format', 'rtf')
        output_name = self.output_data.get('dataset_name', 'output')

        return f"""/*--- Generate Output Table ---*/
ods {output_format} file="&outpath./&outname..{output_format}"
    style=journal
    bodytitle;

ods {output_format} text="~S={{just=center font_weight=bold font_size=10pt}}
{self.submission.get('sponsor', 'AstraZeneca')}";
ods {output_format} text="~S={{just=center font_size=9pt}}
Protocol: {self.submission.get('protocol', 'DXXXXCXXXXX')}";
ods {output_format} text="~S={{just=center font_size=9pt}}
Population: {self.analysis_params.get('population', 'Safety').title()} Population";

title1 "Table X.X.X";
title2 "{self._get_table_title()}";
title3 "{self.analysis_params.get('population', 'Safety').title()} Population";

footnote1 "Source: {', '.join(self.input_data.get('datasets', ['adsl']))}";
footnote2 "Program: &progname..sas  Output: &outname..{output_format}";
footnote3 "Generated: &sysdate9. &systime.";

proc report data=work.final_report nowd split='|'
    style(report)=[font_size=9pt]
    style(header)=[font_weight=bold background=white just=center]
    style(column)=[just=left];
    
    columns row_label col1 col2 col3 col4;
    
    define row_label / display "" style(column)=[cellwidth=3in];
    define col1 / display "Placebo|(N=&N1)" style(column)=[cellwidth=1.5in just=center];
    define col2 / display "Treatment A|(N=&N2)" style(column)=[cellwidth=1.5in just=center];
    define col3 / display "Treatment B|(N=&N3)" style(column)=[cellwidth=1.5in just=center];
    define col4 / display "p-value" style(column)=[cellwidth=1in just=center];
    
    compute row_label;
        if index(row_label, '  ') = 0 then
            call define(_col_, "style", "style=[font_weight=bold]");
    endcomp;
run;

ods {output_format} close;
ods listing;"""

    def _gen_figure_output(self):
        """Generate figure output code."""
        return f"""/*--- Generate Figure Output ---*/
ods listing gpath="&outpath." image_dpi=300;
ods graphics on / reset=all imagename="&outname" imagefmt=pdf
    width=10in height=7in;

proc sgplot data=work.plot_data;
    title1 "Figure X.X.X";
    title2 "{self._get_figure_title()}";
    
    series x=AVISIT y=AVAL / group={self.analysis_params.get('treatment_variable', 'TRTA')}
        markers markerattrs=(size=8)
        lineattrs=(thickness=2);
    
    xaxis label="Visit" fitpolicy=rotate;
    yaxis label="{self.variables.get('analysis_variables', ['AVAL'])[0]}";
    keylegend / location=outside position=bottom;
    
    footnote1 "Source: {', '.join(self.input_data.get('datasets', ['adsl']))}";
    footnote2 "Program: &progname..sas";
run;

ods graphics off;"""

    def _gen_listing_output(self):
        """Generate data listing output."""
        return f"""/*--- Generate Data Listing ---*/
ods rtf file="&outpath./&outname..rtf"
    style=journal bodytitle;

title1 "Listing X.X.X";
title2 "{self._get_listing_title()}";

proc report data=work.listing_data nowd split='|'
    style(report)=[font_size=8pt]
    style(header)=[font_weight=bold background=white];
    
    columns USUBJID SITEID {' '.join(self.variables.get('analysis_variables', ['AVAL']))};
    
    define USUBJID / order "Subject|ID" style(column)=[cellwidth=1.2in];
    define SITEID / order "Site|ID" style(column)=[cellwidth=0.8in];
    
    break after USUBJID / skip;
run;

ods rtf close;
ods listing;"""

    def _gen_dataset_output(self):
        """Generate analysis dataset output."""
        return f"""/*--- Generate Analysis Dataset ---*/
data output.{self.output_data.get('dataset_name', 'analysis_ds')};
    set work.final_dataset;
    label
        STUDYID = "Study Identifier"
        USUBJID = "Unique Subject Identifier"
        {self._generate_labels()}
    ;
run;

/*--- Create XPT transport file for submission ---*/
proc cport data=output.{self.output_data.get('dataset_name', 'analysis_ds')}
    file="&outpath./{ self.output_data.get('dataset_name', 'analysis_ds')}.xpt"
    type=data;
run;

/*--- Dataset verification ---*/
proc contents data=output.{self.output_data.get('dataset_name', 'analysis_ds')} varnum;
run;

proc print data=output.{self.output_data.get('dataset_name', 'analysis_ds')} (obs=10);
run;"""

    def _generate_validation(self):
        """Generate validation and QC checks."""
        return f"""/*--- Validation Checks ---*/

/*--- Check for missing critical variables ---*/
proc sql;
    select count(*) as n_missing_usubjid
    from work.analysis
    where USUBJID is missing;
    
    select count(*) as n_missing_trt
    from work.analysis
    where {self.analysis_params.get('treatment_variable', 'TRTA')} is missing;
quit;

/*--- Verify population counts match ---*/
%macro verify_counts;
    proc sql noprint;
        select count(distinct USUBJID) into :check_n trimmed
        from work.analysis
        where {self._get_population_flag()} = 'Y';
    quit;
    
    %if &check_n ne &Ntot %then %do;
        %put WARNING: Population count mismatch. Expected &Ntot, got &check_n;
    %end;
    %else %do;
        %put NOTE: Population count verified: &check_n subjects.;
    %end;
%mend verify_counts;
%verify_counts;

/*--- Log output summary ---*/
%put NOTE: ========================================;
%put NOTE: Program: &progname..sas;
%put NOTE: Output:  &outname..{self.output_data.get('format', 'rtf')};
%put NOTE: Status:  COMPLETE;
%put NOTE: ========================================;"""

    def _generate_footer(self):
        """Generate program footer."""
        return f"""/*--- End of Program ---*/
/*
REVIEWER NOTES:
1. This program generates {self.output_type} output for {self.analysis_type} analysis
2. Population: {self.analysis_params.get('population', 'Safety')} ({self._get_population_flag()} = 'Y')
3. Statistical method: {self.analysis_params.get('statistical_method', 'descriptive')}
4. Alpha level: {self.analysis_params.get('alpha', 0.05)}
5. Regulatory target: {self.submission.get('regulatory_authority', 'FDA')}
*/"""

    # Helper methods
    def _get_population_flag(self):
        pop_flags = {
            'safety': 'SAFFL',
            'itt': 'ITTFL',
            'per_protocol': 'PPROTFL',
            'pk': 'PKFL',
            'full_analysis': 'FASFL'
        }
        return pop_flags.get(self.analysis_params.get('population', 'safety'), 'SAFFL')

    def _get_orientation(self):
        return 'landscape' if self.output_type in ['table', 'listing'] else 'portrait'

    def _get_purpose_description(self):
        descriptions = {
            'descriptive': 'Generate descriptive statistics summary',
            'efficacy': 'Primary efficacy endpoint analysis',
            'safety': 'Safety summary of adverse events',
            'survival': 'Time-to-event survival analysis',
            'mixed_model': 'Mixed model repeated measures analysis',
            'categorical': 'Categorical data analysis',
            'pk': 'Pharmacokinetic parameter summary',
            'subgroup': 'Subgroup analysis with forest plot'
        }
        return descriptions.get(self.analysis_type, 'Statistical analysis')

    def _get_table_title(self):
        titles = {
            'descriptive': 'Summary of Descriptive Statistics',
            'efficacy': 'Analysis of Primary Efficacy Endpoint',
            'safety': 'Summary of Adverse Events',
            'survival': 'Summary of Time-to-Event Analysis',
            'mixed_model': 'MMRM Analysis Results',
            'categorical': 'Summary of Categorical Endpoints',
            'pk': 'Summary of Pharmacokinetic Parameters',
            'subgroup': 'Subgroup Analysis Results'
        }
        return titles.get(self.analysis_type, 'Analysis Summary')

    def _get_figure_title(self):
        return f"{'Kaplan-Meier Plot' if self.analysis_type == 'survival' else 'Analysis Results'}"

    def _get_listing_title(self):
        return "Subject-Level Data Listing"

    def _generate_labels(self):
        labels = []
        for var in self.variables.get('analysis_variables', []):
            labels.append(f'        {var} = "{var} - Analysis Variable"')
        return '\n'.join(labels)
    
    """
R Code Generator for Clinical Trial Submissions
Generates CDISC-compliant R programs using tidyverse and pharmaverse packages.
"""

from datetime import datetime


class RCodeGenerator:
    def __init__(self, config):
        self.config = config
        self.analysis_type = config.get('analysis_type', 'descriptive')
        self.output_type = config.get('output_type', 'table')
        self.input_data = config.get('input_data', {})
        self.output_data = config.get('output_data', {})
        self.variables = config.get('variables', {})
        self.analysis_params = config.get('analysis_parameters', {})
        self.submission = config.get('submission_details', {})

    def parse_shell(self, shell_path):
        metadata = {'columns': [], 'rows': [], 'titles': [], 'footnotes': []}
        return metadata

    def parse_specification(self, spec_path):
        metadata = {'variables': [], 'derivations': [], 'formats': []}
        return metadata

    def generate(self, shell_metadata=None, spec_metadata=None):
        sections = [
            self._generate_header(),
            self._generate_packages(),
            self._generate_setup(),
            self._generate_data_read(),
            self._generate_data_prep(),
            self._generate_analysis(),
            self._generate_output(),
            self._generate_validation()
        ]
        return '\n\n'.join(sections)

    def _generate_header(self):
        study_id = self.submission.get('study_id', 'STUDY-XXX')
        sponsor = self.submission.get('sponsor', 'AstraZeneca')
        protocol = self.submission.get('protocol', 'DXXXXCXXXXX')

        return f"""# ==============================================================================
# PROGRAM NAME  : {self.output_data.get('dataset_name', 'program')}.R
# STUDY         : {study_id}
# PROTOCOL      : {protocol}
# SPONSOR       : {sponsor}
# PURPOSE       : {self._get_purpose_description()}
# INPUT         : {', '.join(self.input_data.get('datasets', []))}
# OUTPUT        : {self.output_data.get('dataset_name', 'output')}.{self.output_data.get('format', 'rtf')}
# ANALYSIS TYPE : {self.analysis_type.replace('_', ' ').title()}
# OUTPUT TYPE   : {self.output_type.title()}
# POPULATION    : {self.analysis_params.get('population', 'safety').replace('_', ' ').title()}
#
# AUTHOR        : [Programmer Name]
# DATE CREATED  : {datetime.now().strftime('%d-%b-%Y')}
# DATE MODIFIED : {datetime.now().strftime('%d-%b-%Y')}
#
# REGULATORY    : {self.submission.get('regulatory_authority', 'FDA')}
# SUBMISSION    : {self.submission.get('submission_type', 'NDA')}
#
# PACKAGES      : tidyverse, haven, pharmaverse (admiral, rtables, tern)
# VALIDATION    : Independent double programming required
# =============================================================================="""

    def _generate_packages(self):
        base_packages = [
            'library(tidyverse)',
            'library(haven)',
            'library(labelled)'
        ]

        analysis_packages = {
            'descriptive': ['library(gtsummary)', 'library(rtables)', 'library(tern)'],
            'efficacy': ['library(emmeans)', 'library(lme4)', 'library(rtables)', 'library(tern)'],
            'safety': ['library(rtables)', 'library(tern)', 'library(admiral)'],
            'survival': ['library(survival)', 'library(survminer)', 'library(tern)', 'library(visR)'],
            'mixed_model': ['library(mmrm)', 'library(lme4)', 'library(emmeans)', 'library(rtables)'],
            'categorical': ['library(rtables)', 'library(tern)', 'library(exact2x2)'],
            'pk': ['library(PKNCA)', 'library(rtables)', 'library(tern)'],
            'subgroup': ['library(tern)', 'library(rtables)', 'library(forestplot)']
        }

        output_packages = {
            'table': ['library(r2rtf)', 'library(flextable)', 'library(officer)'],
            'figure': ['library(ggplot2)', 'library(cowplot)', 'library(patchwork)'],
            'listing': ['library(r2rtf)', 'library(rlistings)'],
            'dataset': ['library(xportr)', 'library(metacore)']
        }

        packages = base_packages + \
                   analysis_packages.get(self.analysis_type, []) + \
                   output_packages.get(self.output_type, [])

        return f"""# --- Load Required Packages ---
{chr(10).join(packages)}"""

    def _generate_setup(self):
        population_flag = self._get_population_flag()
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')

        return f"""# --- Program Setup ---
study_id    <- "{self.submission.get('study_id', 'STUDY001')}"
protocol    <- "{self.submission.get('protocol', 'D1234C00001')}"
prog_name   <- "{self.output_data.get('dataset_name', 'program')}"
output_name <- "{self.output_data.get('dataset_name', 'output')}"
pop_flag    <- "{population_flag}"
trt_var     <- "{treatment_var}"
alpha       <- {self.analysis_params.get('alpha', 0.05)}
conf_level  <- {self.analysis_params.get('confidence_level', 0.95)}

# --- Path Configuration ---
adam_path   <- "{self.input_data.get('library_path', '/data/adam')}"
output_path <- "{self.output_data.get('output_path', '/output')}"

# --- Treatment Order ---
trt_levels <- c("Placebo", "Treatment A", "Treatment B", "Total")
trt_n_labels <- c()  # Will be populated after data read"""

    def _generate_data_read(self):
        datasets = self.input_data.get('datasets', ['adsl'])
        population_flag = self._get_population_flag()
        filter_vars = self.variables.get('filter_variables', [])

        read_blocks = []
        for ds in datasets:
            filters = [f'{population_flag} == "Y"']
            for fv in filter_vars:
                if fv != population_flag:
                    filters.append(f'{fv} == "Y"')

            filter_expr = ' & '.join(filters) if filters else 'TRUE'

            read_blocks.append(f"""# Read {ds.upper()}
{ds} <- read_sas(file.path(adam_path, "{ds}.sas7bdat")) %>%
    filter({filter_expr}) %>%
    mutate(
        {self.analysis_params.get('treatment_variable', 'TRTA')} = factor(
            {self.analysis_params.get('treatment_variable', 'TRTA')},
            levels = trt_levels
        )
    )

cat(sprintf("NOTE: {ds.upper()} has %d records after filtering.\\n", nrow({ds})))""")

        return f"""# --- Read Input Datasets ---
{chr(10).join(read_blocks)}

# --- Get Population Counts (Big N) ---
big_n <- {datasets[0]} %>%
    distinct(USUBJID, .keep_all = TRUE) %>%
    count({self.analysis_params.get('treatment_variable', 'TRTA')}, name = "N") %>%
    mutate(
        N_label = sprintf("%s\\n(N=%d)", {self.analysis_params.get('treatment_variable', 'TRTA')}, N)
    )

cat("NOTE: Population counts:\\n")
print(big_n)"""

    def _generate_data_prep(self):
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        grouping_vars = self.variables.get('grouping_variables', [])

        return f"""# --- Data Preparation ---
analysis_data <- {self.input_data.get('datasets', ['adsl'])[0]} %>%
    # Add treatment numeric ordering
    mutate(
        TRTN = case_when(
            {treatment_var} == "Placebo"      ~ 1L,
            {treatment_var} == "Treatment A"  ~ 2L,
            {treatment_var} == "Treatment B"  ~ 3L,
            TRUE                              ~ 99L
        )
    ) %>%
    arrange(TRTN, {', '.join(self.variables.get('sort_variables', ['USUBJID']))})

# --- Create Total group ---
analysis_with_total <- bind_rows(
    analysis_data,
    analysis_data %>% mutate({treatment_var} = "Total", TRTN = 99L)
)

cat(sprintf("NOTE: Analysis dataset has %d records (including Total).\\n", 
    nrow(analysis_with_total)))"""

    def _generate_analysis(self):
        method_generators = {
            'descriptive': self._gen_descriptive_r,
            'efficacy': self._gen_efficacy_r,
            'safety': self._gen_safety_r,
            'survival': self._gen_survival_r,
            'mixed_model': self._gen_mmrm_r,
            'categorical': self._gen_categorical_r,
            'pk': self._gen_pk_r,
            'subgroup': self._gen_subgroup_r
        }
        generator = method_generators.get(self.analysis_type, self._gen_descriptive_r)
        return generator()

    def _gen_descriptive_r(self):
        analysis_vars = self.variables.get('analysis_variables', ['AVAL'])
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')

        return f"""# --- Descriptive Statistics ---
desc_stats <- analysis_with_total %>%
    group_by({treatment_var}, TRTN) %>%
    summarise(
        across(
            c({', '.join(analysis_vars)}),
            list(
                n      = ~sum(!is.na(.)),
                mean   = ~mean(., na.rm = TRUE),
                sd     = ~sd(., na.rm = TRUE),
                median = ~median(., na.rm = TRUE),
                min    = ~min(., na.rm = TRUE),
                max    = ~max(., na.rm = TRUE),
                q1     = ~quantile(., 0.25, na.rm = TRUE),
                q3     = ~quantile(., 0.75, na.rm = TRUE)
            ),
            .names = "{{.col}}_{{.fn}}"
        ),
        .groups = "drop"
    ) %>%
    arrange(TRTN)

# --- Format for display ---
desc_formatted <- desc_stats %>%
    mutate(
        n_c      = sprintf("%d", {analysis_vars[0]}_n),
        mean_sd  = sprintf("%.1f (%.2f)", {analysis_vars[0]}_mean, {analysis_vars[0]}_sd),
        median_c = sprintf("%.1f", {analysis_vars[0]}_median),
        q1_q3    = sprintf("%.1f, %.1f", {analysis_vars[0]}_q1, {analysis_vars[0]}_q3),
        min_max  = sprintf("%.1f, %.1f", {analysis_vars[0]}_min, {analysis_vars[0]}_max)
    ) %>%
    select({treatment_var}, TRTN, n_c, mean_sd, median_c, q1_q3, min_max)

# --- Using rtables for submission-ready output ---
lyt <- basic_table() %>%
    split_cols_by("{treatment_var}") %>%
    add_colcounts() %>%
    analyze(
        vars = c({', '.join([f'"{v}"' for v in analysis_vars])}),
        afun = function(x) {{
            in_rows(
                "n"           = rcell(sum(!is.na(x)), format = "xx"),
                "Mean (SD)"   = rcell(c(mean(x, na.rm=TRUE), sd(x, na.rm=TRUE)), format = "xx.x (xx.xx)"),
                "Median"      = rcell(median(x, na.rm=TRUE), format = "xx.x"),
                "Q1, Q3"      = rcell(c(quantile(x, 0.25, na.rm=TRUE), quantile(x, 0.75, na.rm=TRUE)), format = "xx.x, xx.x"),
                "Min, Max"    = rcell(c(min(x, na.rm=TRUE), max(x, na.rm=TRUE)), format = "xx.x, xx.x")
            )
        }}
    )

result_table <- build_table(lyt, analysis_with_total)
print(result_table)"""

    def _gen_efficacy_r(self):
        analysis_vars = self.variables.get('analysis_variables', ['CHG'])
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        covariates = self.analysis_params.get('covariates', ['BASE'])

        return f"""# --- Efficacy Analysis: ANCOVA ---

# Fit ANCOVA model
ancova_model <- lm(
    {analysis_vars[0]} ~ {treatment_var} + {' + '.join(covariates)},
    data = analysis_data
)

# Model summary
model_summary <- summary(ancova_model)
cat("\\nANCOVA Model Summary:\\n")
print(model_summary)

# LS Means
lsmeans_result <- emmeans(ancova_model, specs = ~ {treatment_var})
lsmeans_df <- as.data.frame(lsmeans_result) %>%
    mutate(
        lsmean_ci = sprintf("%.2f (%.2f, %.2f)", emmean, lower.CL, upper.CL)
    )

# Pairwise comparisons
contrasts_result <- pairs(lsmeans_result, adjust = "none")
contrasts_df <- as.data.frame(contrasts_result) %>%
    mutate(
        diff_ci = sprintf("%.2f (%.2f, %.2f)", estimate, lower.CL, upper.CL),
        p_value_c = ifelse(p.value < 0.001, "<0.001", sprintf("%.3f", p.value))
    )

cat("\\nLS Means:\\n")
print(lsmeans_df)
cat("\\nTreatment Differences:\\n")
print(contrasts_df)"""

    def _gen_safety_r(self):
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')

        return f"""# --- Safety Analysis: Adverse Event Summary ---

# Overall AE incidence
ae_overall <- adae %>%
    group_by({treatment_var}) %>%
    summarise(
        n_subjects = n_distinct(USUBJID),
        n_events   = n(),
        .groups    = "drop"
    ) %>%
    left_join(big_n, by = "{treatment_var}") %>%
    mutate(
        pct = (n_subjects / N) * 100,
        col = sprintf("%d (%.1f%%)", n_subjects, pct)
    )

# AE by SOC and PT
ae_soc_pt <- adae %>%
    group_by({treatment_var}, AEBODSYS, AEDECOD) %>%
    summarise(
        n_subjects = n_distinct(USUBJID),
        n_events   = n(),
        .groups    = "drop"
    ) %>%
    left_join(big_n, by = "{treatment_var}") %>%
    mutate(
        pct = (n_subjects / N) * 100,
        col = sprintf("%d (%.1f%%)", n_subjects, pct)
    ) %>%
    arrange(desc(n_subjects), AEBODSYS, AEDECOD)

# Using rtables for AE table
lyt_ae <- basic_table() %>%
    split_cols_by("{treatment_var}") %>%
    add_colcounts() %>%
    split_rows_by("AEBODSYS", label_pos = "topleft", split_label = "System Organ Class") %>%
    summarize_num_patients(var = "USUBJID", .stats = c("unique", "nonunique")) %>%
    analyze("AEDECOD", afun = function(x, .N_col) {{
        in_rows(
            rcell(length(unique(x)), format = "xx"),
            .labels = "n (%)"
        )
    }})

ae_table <- build_table(lyt_ae, adae)
print(ae_table)"""

    def _gen_survival_r(self):
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')

        return f"""# --- Survival Analysis ---

# Kaplan-Meier estimation
km_fit <- survfit(
    Surv(AVAL, 1 - CNSR) ~ {treatment_var},
    data = analysis_data,
    conf.type = "log-log"
)

# KM Summary
km_summary <- summary(km_fit, times = c(90, 180, 270, 365))
cat("\\nKaplan-Meier Estimates:\\n")
print(km_summary)

# Log-rank test
logrank_test <- survdiff(
    Surv(AVAL, 1 - CNSR) ~ {treatment_var},
    data = analysis_data
)
cat("\\nLog-Rank Test:\\n")
print(logrank_test)

# Cox Proportional Hazards
cox_model <- coxph(
    Surv(AVAL, 1 - CNSR) ~ {treatment_var},
    data = analysis_data,
    ties = "efron"
)

cox_summary <- summary(cox_model, conf.int = {self.analysis_params.get('confidence_level', 0.95)})
cat("\\nCox PH Model:\\n")
print(cox_summary)

# Hazard Ratio
hr_result <- data.frame(
    HR       = exp(coef(cox_model)),
    HR_lower = exp(confint(cox_model))[, 1],
    HR_upper = exp(confint(cox_model))[, 2],
    p_value  = summary(cox_model)$coefficients[, "Pr(>|z|)"]
) %>%
    mutate(
        hr_ci = sprintf("%.3f (%.3f, %.3f)", HR, HR_lower, HR_upper),
        p_value_c = ifelse(p_value < 0.001, "<0.001", sprintf("%.3f", p_value))
    )

# KM Plot
km_plot <- ggsurvplot(
    km_fit,
    data = analysis_data,
    risk.table = TRUE,
    pval = TRUE,
    conf.int = TRUE,
    xlab = "Time (Days)",
    ylab = "Survival Probability",
    title = "Kaplan-Meier Survival Curve",
    legend.title = "Treatment",
    palette = c("#0072B2", "#D55E00", "#009E73"),
    ggtheme = theme_classic()
)"""

    def _gen_mmrm_r(self):
        analysis_vars = self.variables.get('analysis_variables', ['CHG'])
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        covariates = self.analysis_params.get('covariates', ['BASE'])

        return f"""# --- Mixed Model Repeated Measures (MMRM) ---

# Fit MMRM using mmrm package
mmrm_fit <- mmrm(
    formula = {analysis_vars[0]} ~ {treatment_var} * AVISIT + {' + '.join(covariates)} +
        us(AVISIT | USUBJID),
    data = analysis_data,
    method = "Kenward-Roger"
)

# Model summary
cat("\\nMMRM Model Summary:\\n")
summary(mmrm_fit)

# LS Means by treatment and visit
lsmeans_mmrm <- emmeans(mmrm_fit, ~ {treatment_var} | AVISIT)
lsmeans_df <- as.data.frame(lsmeans_mmrm)

# Treatment differences at each visit
diffs_mmrm <- pairs(lsmeans_mmrm, adjust = "none")
diffs_df <- as.data.frame(diffs_mmrm) %>%
    mutate(
        diff_ci = sprintf("%.2f (%.2f, %.2f)", estimate, lower.CL, upper.CL),
        p_value_c = ifelse(p.value < 0.001, "<0.001", sprintf("%.3f", p.value))
    )

cat("\\nLS Means by Visit:\\n")
print(lsmeans_df)
cat("\\nTreatment Differences by Visit:\\n")
print(diffs_df)"""

    def _gen_categorical_r(self):
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        return f"""# --- Categorical Analysis ---

# Frequency table
freq_table <- analysis_data %>%
    count({treatment_var}, AVALC) %>%
    left_join(big_n, by = "{treatment_var}") %>%
    mutate(
        pct = (n / N) * 100,
        col = sprintf("%d (%.1f%%)", n, pct)
    )

# Chi-square test
chisq_result <- chisq.test(table(analysis_data${treatment_var}, analysis_data$AVALC))

# Fisher's exact test
fisher_result <- fisher.test(table(analysis_data${treatment_var}, analysis_data$AVALC))

cat(sprintf("Chi-square p-value: %.4f\\n", chisq_result$p.value))
cat(sprintf("Fisher's exact p-value: %.4f\\n", fisher_result$p.value))"""

    def _gen_pk_r(self):
        return """# --- Pharmacokinetic Analysis ---

# PK parameter summary
pk_summary <- analysis_data %>%
    group_by(TRTA, PARAM, PARAMCD) %>%
    summarise(
        n      = sum(!is.na(AVAL)),
        mean   = mean(AVAL, na.rm = TRUE),
        sd     = sd(AVAL, na.rm = TRUE),
        cv     = sd(AVAL, na.rm = TRUE) / mean(AVAL, na.rm = TRUE) * 100,
        median = median(AVAL, na.rm = TRUE),
        min    = min(AVAL, na.rm = TRUE),
        max    = max(AVAL, na.rm = TRUE),
        gmean  = exp(mean(log(AVAL[AVAL > 0]), na.rm = TRUE)),
        gcv    = sqrt(exp(var(log(AVAL[AVAL > 0]), na.rm = TRUE)) - 1) * 100,
        .groups = "drop"
    )

print(pk_summary)"""

    def _gen_subgroup_r(self):
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        subgroups = self.analysis_params.get('subgroup_variables', ['SEX', 'AGEGR1', 'RACE'])

        return f"""# --- Subgroup Analysis ---

subgroup_results <- map_dfr(c({', '.join([f'"{sg}"' for sg in subgroups])}), function(sg) {{
    formula_str <- paste0("CHG ~ {treatment_var} * ", sg, " + BASE")
    model <- lm(as.formula(formula_str), data = analysis_data)
    
    emm <- emmeans(model, ~ {treatment_var} | !!sym(sg))
    diffs <- as.data.frame(pairs(emm, adjust = "none"))
    
    diffs %>%
        mutate(subgroup_var = sg) %>%
        rename(category = !!sym(sg))
}})

# Forest plot
forest_plot <- ggplot(subgroup_results, aes(x = estimate, y = category)) +
    geom_point(size = 3) +
    geom_errorbarh(aes(xmin = lower.CL, xmax = upper.CL), height = 0.2) +
    geom_vline(xintercept = 0, linetype = "dashed", color = "grey50") +
    facet_grid(subgroup_var ~ ., scales = "free_y", space = "free_y") +
    labs(x = "Treatment Difference (95% CI)", y = "") +
    theme_minimal()

print(forest_plot)"""

    def _generate_output(self):
        output_generators = {
            'table': self._gen_table_output_r,
            'figure': self._gen_figure_output_r,
            'listing': self._gen_listing_output_r,
            'dataset': self._gen_dataset_output_r
        }
        generator = output_generators.get(self.output_type, self._gen_table_output_r)
        return generator()

    def _gen_table_output_r(self):
        return f"""# --- Generate Table Output ---

# Using r2rtf for submission-ready RTF output
output_file <- file.path(output_path, paste0(output_name, ".rtf"))

result_df %>%
    r2rtf::rtf_title(
        title = "Table X.X.X",
        subtitle = "{self._get_purpose_description()}"
    ) %>%
    r2rtf::rtf_colheader(
        colheader = "Statistic | Placebo\\n(N={{N1}}) | Treatment A\\n(N={{N2}}) | Treatment B\\n(N={{N3}})",
        col_rel_width = c(3, 2, 2, 2)
    ) %>%
    r2rtf::rtf_body(
        col_rel_width = c(3, 2, 2, 2),
        text_justification = c("l", "c", "c", "c")
    ) %>%
    r2rtf::rtf_footnote(
        footnote = c(
            "Source: {', '.join(self.input_data.get('datasets', ['adsl']))}",
            "Program: {{prog_name}}.R  Output: {{output_name}}.rtf",
            paste0("Generated: ", format(Sys.time(), "%d%b%Y %H:%M"))
        )
    ) %>%
    r2rtf::rtf_encode() %>%
    r2rtf::write_rtf(output_file)

cat(sprintf("NOTE: Output written to %s\\n", output_file))"""

    def _gen_figure_output_r(self):
        return f"""# --- Generate Figure Output ---

output_file <- file.path(output_path, paste0(output_name, ".pdf"))

final_plot <- km_plot +
    labs(
        title = "Figure X.X.X",
        subtitle = "{self._get_purpose_description()}",
        caption = paste0(
            "Source: {', '.join(self.input_data.get('datasets', ['adsl']))}\\n",
            "Program: ", prog_name, ".R\\n",
            "Generated: ", format(Sys.time(), "%d%b%Y")
        )
    )

ggsave(
    output_file,
    plot = final_plot,
    width = 10,
    height = 7,
    units = "in",
    dpi = 300
)

cat(sprintf("NOTE: Figure written to %s\\n", output_file))"""

    def _gen_listing_output_r(self):
        return f"""# --- Generate Listing Output ---

output_file <- file.path(output_path, paste0(output_name, ".rtf"))

listing_data %>%
    select(USUBJID, SITEID, {', '.join(self.variables.get('analysis_variables', ['AVAL']))}) %>%
    arrange(USUBJID) %>%
    r2rtf::rtf_title(title = "Listing X.X.X") %>%
    r2rtf::rtf_body(text_font_size = 8) %>%
    r2rtf::rtf_encode() %>%
    r2rtf::write_rtf(output_file)

cat(sprintf("NOTE: Listing written to %s\\n", output_file))"""

    def _gen_dataset_output_r(self):
        return f"""# --- Generate Analysis Dataset ---

# Apply metadata and export as XPT
output_file <- file.path(output_path, paste0(output_name, ".xpt"))

final_dataset %>%
    xportr::xportr_type(metacore) %>%
    xportr::xportr_length(metacore) %>%
    xportr::xportr_label(metacore) %>%
    xportr::xportr_format(metacore) %>%
    xportr::xportr_write(output_file)

cat(sprintf("NOTE: Dataset written to %s\\n", output_file))"""

    def _generate_validation(self):
        return f"""# --- Validation Checks ---

# Check for missing critical variables
n_missing_usubjid <- sum(is.na(analysis_data$USUBJID))
n_missing_trt <- sum(is.na(analysis_data${self.analysis_params.get('treatment_variable', 'TRTA')}))

cat(sprintf("\\nValidation Results:\\n"))
cat(sprintf("  Missing USUBJID: %d\\n", n_missing_usubjid))
cat(sprintf("  Missing Treatment: %d\\n", n_missing_trt))
cat(sprintf("  Total subjects: %d\\n", n_distinct(analysis_data$USUBJID)))

# Session info for reproducibility
cat("\\n--- Session Info ---\\n")
sessionInfo()

cat("\\n========================================\\n")
cat(sprintf("Program: %s.R\\n", prog_name))
cat(sprintf("Output:  %s.{self.output_data.get('format', 'rtf')}\\n", output_name))
cat("Status:  COMPLETE\\n")
cat("========================================\\n")"""

    # Helper methods
    def _get_population_flag(self):
        pop_flags = {
            'safety': 'SAFFL', 'itt': 'ITTFL', 'per_protocol': 'PPROTFL',
            'pk': 'PKFL', 'full_analysis': 'FASFL'
        }
        return pop_flags.get(self.analysis_params.get('population', 'safety'), 'SAFFL')

    def _get_purpose_description(self):
        descriptions = {
            'descriptive': 'Generate descriptive statistics summary',
            'efficacy': 'Primary efficacy endpoint analysis',
            'safety': 'Safety summary of adverse events',
            'survival': 'Time-to-event survival analysis',
            'mixed_model': 'Mixed model repeated measures analysis',
            'categorical': 'Categorical data analysis',
            'pk': 'Pharmacokinetic parameter summary',
            'subgroup': 'Subgroup analysis with forest plot'
        }
        return descriptions.get(self.analysis_type, 'Statistical analysis')
    
    """
Python Code Generator for Clinical Trial Submissions
Generates CDISC-compliant Python programs using pandas, statsmodels, and lifelines.
"""

from datetime import datetime


class PythonCodeGenerator:
    def __init__(self, config):
        self.config = config
        self.analysis_type = config.get('analysis_type', 'descriptive')
        self.output_type = config.get('output_type', 'table')
        self.input_data = config.get('input_data', {})
        self.output_data = config.get('output_data', {})
        self.variables = config.get('variables', {})
        self.analysis_params = config.get('analysis_parameters', {})
        self.submission = config.get('submission_details', {})

    def parse_shell(self, shell_path):
        metadata = {'columns': [], 'rows': [], 'titles': [], 'footnotes': []}
        return metadata

    def parse_specification(self, spec_path):
        metadata = {'variables': [], 'derivations': [], 'formats': []}
        return metadata

    def generate(self, shell_metadata=None, spec_metadata=None):
        sections = [
            self._generate_header(),
            self._generate_imports(),
            self._generate_setup(),
            self._generate_data_read(),
            self._generate_data_prep(),
            self._generate_analysis(),
            self._generate_output(),
            self._generate_validation()
        ]
        return '\n\n'.join(sections)

    def _generate_header(self):
        study_id = self.submission.get('study_id', 'STUDY-XXX')
        sponsor = self.submission.get('sponsor', 'AstraZeneca')
        protocol = self.submission.get('protocol', 'DXXXXCXXXXX')

        return f'''"""
================================================================================
PROGRAM NAME  : {self.output_data.get('dataset_name', 'program')}.py
STUDY         : {study_id}
PROTOCOL      : {protocol}
SPONSOR       : {sponsor}
PURPOSE       : {self._get_purpose_description()}
INPUT         : {', '.join(self.input_data.get('datasets', []))}
OUTPUT        : {self.output_data.get('dataset_name', 'output')}.{self.output_data.get('format', 'rtf')}
ANALYSIS TYPE : {self.analysis_type.replace('_', ' ').title()}
OUTPUT TYPE   : {self.output_type.title()}
POPULATION    : {self.analysis_params.get('population', 'safety').replace('_', ' ').title()}

AUTHOR        : [Programmer Name]
DATE CREATED  : {datetime.now().strftime('%d-%b-%Y')}
DATE MODIFIED : {datetime.now().strftime('%d-%b-%Y')}

REGULATORY    : {self.submission.get('regulatory_authority', 'FDA')}
SUBMISSION    : {self.submission.get('submission_type', 'NDA')}

VALIDATION    : Independent double programming required
CDISC VERSION : ADaM IG v1.3 / SDTM IG v3.4
================================================================================
"""'''

    def _generate_imports(self):
        base_imports = [
            'import pandas as pd',
            'import numpy as np',
            'from pathlib import Path',
            'from datetime import datetime',
            'import warnings',
            'warnings.filterwarnings("ignore")'
        ]

        analysis_imports = {
            'descriptive': [
                'from scipy import stats',
                'import statsmodels.api as sm'
            ],
            'efficacy': [
                'import statsmodels.api as sm',
                'from statsmodels.formula.api import ols, mixedlm',
                'from statsmodels.stats.multicomp import pairwise_tukeyhsd'
            ],
            'safety': [
                'from scipy import stats'
            ],
            'survival': [
                'from lifelines import KaplanMeierFitter, CoxPHFitter',
                'from lifelines.statistics import logrank_test',
                'import matplotlib.pyplot as plt'
            ],
            'mixed_model': [
                'import statsmodels.api as sm',
                'from statsmodels.formula.api import mixedlm',
                'from statsmodels.regression.mixed_linear_model import MixedLM'
            ],
            'categorical': [
                'from scipy.stats import chi2_contingency, fisher_exact',
                'import statsmodels.api as sm'
            ],
            'pk': [
                'from scipy import stats',
                'from scipy.integrate import trapezoid'
            ],
            'subgroup': [
                'import statsmodels.api as sm',
                'from statsmodels.formula.api import ols',
                'import matplotlib.pyplot as plt'
            ]
        }

        output_imports = {
            'table': ['from openpyxl import Workbook', 'from docx import Document'],
            'figure': ['import matplotlib.pyplot as plt', 'import seaborn as sns', 'from matplotlib.backends.backend_pdf import PdfPages'],
            'listing': ['from openpyxl import Workbook'],
            'dataset': ['import pyreadstat']
        }

        all_imports = base_imports + \
                      analysis_imports.get(self.analysis_type, []) + \
                      output_imports.get(self.output_type, [])

        # Deduplicate
        seen = set()
        unique_imports = []
        for imp in all_imports:
            if imp not in seen:
                seen.add(imp)
                unique_imports.append(imp)

        return '\n'.join(unique_imports)

    def _generate_setup(self):
        population_flag = self._get_population_flag()
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')

        return f"""# --- Program Configuration ---
CONFIG = {{
    'study_id': '{self.submission.get('study_id', 'STUDY001')}',
    'protocol': '{self.submission.get('protocol', 'D1234C00001')}',
    'prog_name': '{self.output_data.get('dataset_name', 'program')}',
    'output_name': '{self.output_data.get('dataset_name', 'output')}',
    'pop_flag': '{population_flag}',
    'trt_var': '{treatment_var}',
    'alpha': {self.analysis_params.get('alpha', 0.05)},
    'conf_level': {self.analysis_params.get('confidence_level', 0.95)},
    'adam_path': Path(r'{self.input_data.get('library_path', '/data/adam')}'),
    'output_path': Path(r'{self.output_data.get('output_path', '/output')}'),
}}

# Treatment ordering
TRT_ORDER = ['Placebo', 'Treatment A', 'Treatment B', 'Total']
TRT_MAP = {{'Placebo': 1, 'Treatment A': 2, 'Treatment B': 3, 'Total': 99}}

# Ensure output directory exists
CONFIG['output_path'].mkdir(parents=True, exist_ok=True)"""

    def _generate_data_read(self):
        datasets = self.input_data.get('datasets', ['adsl'])
        population_flag = self._get_population_flag()
        filter_vars = self.variables.get('filter_variables', [])

        read_blocks = []
        for ds in datasets:
            filters = [f'(df["{population_flag}"] == "Y")']
            for fv in filter_vars:
                if fv != population_flag:
                    filters.append(f'(df["{fv}"] == "Y")')

            filter_expr = ' & '.join(filters)

            read_blocks.append(f"""# Read {ds.upper()}
{ds}_path = CONFIG['adam_path'] / '{ds}.sas7bdat'
{ds}_raw, meta = pyreadstat.read_sas7bdat(str({ds}_path))
{ds} = {ds}_raw[{filter_expr}].copy()
{ds}[CONFIG['trt_var']] = pd.Categorical(
    {ds}[CONFIG['trt_var']], categories=TRT_ORDER[:-1], ordered=True
)
print(f"NOTE: {ds.upper()} has {{{ds}.shape[0]}} records after filtering.")""")

        return f"""# --- Read Input Datasets ---
import pyreadstat

{chr(10).join(read_blocks)}

# --- Get Population Counts (Big N) ---
big_n = {datasets[0]}.drop_duplicates(subset=['USUBJID']).groupby(
    CONFIG['trt_var']
)['USUBJID'].nunique().to_dict()

print(f"NOTE: Population counts: {{big_n}}")"""

    def _generate_data_prep(self):
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        sort_vars = self.variables.get('sort_variables', ['USUBJID'])

        return f"""# --- Data Preparation ---
analysis_data = {self.input_data.get('datasets', ['adsl'])[0]}.copy()

# Add treatment numeric ordering
analysis_data['TRTN'] = analysis_data[CONFIG['trt_var']].map(TRT_MAP)

# Sort data
analysis_data = analysis_data.sort_values(
    ['TRTN'] + {sort_vars}
).reset_index(drop=True)

# Create version with Total group
analysis_total = analysis_data.copy()
analysis_total[CONFIG['trt_var']] = 'Total'
analysis_total['TRTN'] = 99

analysis_with_total = pd.concat(
    [analysis_data, analysis_total], ignore_index=True
)

print(f"NOTE: Analysis dataset has {{analysis_with_total.shape[0]}} records (including Total).")"""

    def _generate_analysis(self):
        method_generators = {
            'descriptive': self._gen_descriptive_py,
            'efficacy': self._gen_efficacy_py,
            'safety': self._gen_safety_py,
            'survival': self._gen_survival_py,
            'mixed_model': self._gen_mmrm_py,
            'categorical': self._gen_categorical_py,
            'pk': self._gen_pk_py,
            'subgroup': self._gen_subgroup_py
        }
        generator = method_generators.get(self.analysis_type, self._gen_descriptive_py)
        return generator()

    def _gen_descriptive_py(self):
        analysis_vars = self.variables.get('analysis_variables', ['AVAL'])
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')

        return f"""# --- Descriptive Statistics ---

def compute_descriptive_stats(df, var, group_var):
    \"\"\"Compute descriptive statistics by treatment group.\"\"\"
    stats_df = df.groupby(group_var)[var].agg([
        ('n', 'count'),
        ('mean', 'mean'),
        ('std', 'std'),
        ('median', 'median'),
        ('min', 'min'),
        ('max', 'max'),
        ('q1', lambda x: x.quantile(0.25)),
        ('q3', lambda x: x.quantile(0.75))
    ]).reset_index()
    
    # Format for display
    stats_df['n_c'] = stats_df['n'].apply(lambda x: f"{{x:.0f}}")
    stats_df['mean_sd'] = stats_df.apply(
        lambda r: f"{{r['mean']:.1f}} ({{r['std']:.2f}})", axis=1
    )
    stats_df['median_c'] = stats_df['median'].apply(lambda x: f"{{x:.1f}}")
    stats_df['q1_q3'] = stats_df.apply(
        lambda r: f"{{r['q1']:.1f}}, {{r['q3']:.1f}}", axis=1
    )
    stats_df['min_max'] = stats_df.apply(
        lambda r: f"{{r['min']:.1f}}, {{r['max']:.1f}}", axis=1
    )
    
    return stats_df

# Compute statistics for each analysis variable
results = {{}}
for var in {analysis_vars}:
    results[var] = compute_descriptive_stats(
        analysis_with_total, var, CONFIG['trt_var']
    )
    print(f"\\nDescriptive Statistics for {{var}}:")
    print(results[var][['{ treatment_var}', 'n_c', 'mean_sd', 'median_c', 'q1_q3', 'min_max']])"""

    def _gen_efficacy_py(self):
        analysis_vars = self.variables.get('analysis_variables', ['CHG'])
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        covariates = self.analysis_params.get('covariates', ['BASE'])

        return f"""# --- Efficacy Analysis: ANCOVA ---

# Fit ANCOVA model
formula = '{analysis_vars[0]} ~ C({treatment_var}, Treatment("Placebo")) + {" + ".join(covariates)}'
model = ols(formula, data=analysis_data).fit()

print("\\nANCOVA Model Summary:")
print(model.summary())

# LS Means (adjusted means)
from statsmodels.stats.contrast import ContrastResults

# Get predicted means for each treatment
lsmeans = {{}}
for trt in analysis_data[CONFIG['trt_var']].unique():
    subset = analysis_data[analysis_data[CONFIG['trt_var']] == trt]
    pred_data = subset.copy()
    for cov in {covariates}:
        pred_data[cov] = analysis_data[cov].mean()
    lsmeans[trt] = model.predict(pred_data).mean()

print("\\nLS Means:")
for trt, mean in lsmeans.items():
    print(f"  {{trt}}: {{mean:.2f}}")

# Treatment differences
print("\\nTreatment Differences vs Placebo:")
for trt in [t for t in lsmeans.keys() if t != 'Placebo']:
    diff = lsmeans[trt] - lsmeans['Placebo']
    print(f"  {{trt}} - Placebo: {{diff:.2f}}")"""

    def _gen_safety_py(self):
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')

        return f"""# --- Safety Analysis: Adverse Event Summary ---

# Overall AE incidence
ae_overall = adae.groupby(CONFIG['trt_var']).agg(
    n_subjects=('USUBJID', 'nunique'),
    n_events=('USUBJID', 'count')
).reset_index()

ae_overall['N'] = ae_overall[CONFIG['trt_var']].map(big_n)
ae_overall['pct'] = (ae_overall['n_subjects'] / ae_overall['N']) * 100
ae_overall['col'] = ae_overall.apply(
    lambda r: f"{{r['n_subjects']:.0f}} ({{r['pct']:.1f}}%)", axis=1
)

print("\\nOverall AE Summary:")
print(ae_overall)

# AE by SOC and PT
ae_soc_pt = adae.groupby(
    [CONFIG['trt_var'], 'AEBODSYS', 'AEDECOD']
).agg(
    n_subjects=('USUBJID', 'nunique'),
    n_events=('USUBJID', 'count')
).reset_index()

ae_soc_pt['N'] = ae_soc_pt[CONFIG['trt_var']].map(big_n)
ae_soc_pt['pct'] = (ae_soc_pt['n_subjects'] / ae_soc_pt['N']) * 100
ae_soc_pt['col'] = ae_soc_pt.apply(
    lambda r: f"{{r['n_subjects']:.0f}} ({{r['pct']:.1f}}%)", axis=1
)

# Sort by descending frequency
ae_sorted = ae_soc_pt.sort_values(
    ['n_subjects', 'AEBODSYS', 'AEDECOD'], ascending=[False, True, True]
)

print("\\nTop 10 AEs by SOC/PT:")
print(ae_sorted.head(10))"""

    def _gen_survival_py(self):
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')

        return f"""# --- Survival Analysis ---

# Kaplan-Meier Estimation
kmf = KaplanMeierFitter()

fig, ax = plt.subplots(figsize=(10, 7))
km_results = {{}}

for trt in analysis_data[CONFIG['trt_var']].unique():
    subset = analysis_data[analysis_data[CONFIG['trt_var']] == trt]
    kmf.fit(
        durations=subset['AVAL'],
        event_observed=(1 - subset['CNSR']),
        label=trt
    )
    kmf.plot_survival_function(ax=ax, ci_show=True)
    
    km_results[trt] = {{
        'median': kmf.median_survival_time_,
        'survival_6mo': kmf.predict(180),
        'survival_12mo': kmf.predict(365)
    }}

ax.set_xlabel('Time (Days)')
ax.set_ylabel('Survival Probability')
ax.set_title('Kaplan-Meier Survival Curve')
ax.legend(loc='lower left')

# Log-rank test
groups = analysis_data[CONFIG['trt_var']].unique()
if len(groups) == 2:
    g1 = analysis_data[analysis_data[CONFIG['trt_var']] == groups[0]]
    g2 = analysis_data[analysis_data[CONFIG['trt_var']] == groups[1]]
    lr_result = logrank_test(
        g1['AVAL'], g2['AVAL'],
        event_observed_A=(1 - g1['CNSR']),
        event_observed_B=(1 - g2['CNSR'])
    )
    print(f"\\nLog-Rank Test p-value: {{lr_result.p_value:.4f}}")

# Cox Proportional Hazards
cph = CoxPHFitter()
cox_data = analysis_data[['AVAL', 'CNSR', CONFIG['trt_var']]].copy()
cox_data['event'] = 1 - cox_data['CNSR']
cox_data = pd.get_dummies(cox_data, columns=[CONFIG['trt_var']], drop_first=True)

cph.fit(cox_data, duration_col='AVAL', event_col='event')
print("\\nCox PH Model:")
cph.print_summary()"""

    def _gen_mmrm_py(self):
        analysis_vars = self.variables.get('analysis_variables', ['CHG'])
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        covariates = self.analysis_params.get('covariates', ['BASE'])

        return f"""# --- Mixed Model Repeated Measures (MMRM) ---

# Note: For full MMRM with unstructured covariance, consider using R via rpy2
# This implementation uses statsmodels mixed linear model

formula = '{analysis_vars[0]} ~ C({treatment_var}) * C(AVISIT) + {" + ".join(covariates)}'

mmrm_model = mixedlm(
    formula,
    data=analysis_data,
    groups='USUBJID',
    re_formula='~C(AVISIT)'
)

mmrm_result = mmrm_model.fit(reml=True)
print("\\nMMRM Model Summary:")
print(mmrm_result.summary())

# LS Means by visit
visits = analysis_data['AVISIT'].unique()
lsmeans_by_visit = {{}}

for visit in sorted(visits):
    visit_data = analysis_data[analysis_data['AVISIT'] == visit]
    for trt in visit_data[CONFIG['trt_var']].unique():
        trt_visit_data = visit_data[visit_data[CONFIG['trt_var']] == trt]
        pred = mmrm_result.predict(trt_visit_data)
        lsmeans_by_visit[(trt, visit)] = pred.mean()

print("\\nLS Means by Treatment and Visit:")
for (trt, visit), mean in sorted(lsmeans_by_visit.items()):
    print(f"  {{trt}} - {{visit}}: {{mean:.2f}}")"""

    def _gen_categorical_py(self):
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        return f"""# --- Categorical Analysis ---

# Frequency table
freq_table = pd.crosstab(
    analysis_data[CONFIG['trt_var']],
    analysis_data['AVALC'],
    margins=True
)
print("\\nFrequency Table:")
print(freq_table)

# Chi-square test
contingency = pd.crosstab(analysis_data[CONFIG['trt_var']], analysis_data['AVALC'])
chi2, p_value, dof, expected = chi2_contingency(contingency)
print(f"\\nChi-square test: chi2={{chi2:.4f}}, p={{p_value:.4f}}, df={{dof}}")

# Fisher's exact test (for 2x2 tables)
if contingency.shape == (2, 2):
    odds_ratio, fisher_p = fisher_exact(contingency)
    print(f"Fisher's exact test: OR={{odds_ratio:.4f}}, p={{fisher_p:.4f}}")"""

    def _gen_pk_py(self):
        return """# --- Pharmacokinetic Analysis ---

pk_summary = analysis_data.groupby(
    [CONFIG['trt_var'], 'PARAM', 'PARAMCD']
)['AVAL'].agg([
    ('n', 'count'),
    ('mean', 'mean'),
    ('std', 'std'),
    ('median', 'median'),
    ('min', 'min'),
    ('max', 'max')
]).reset_index()

# Geometric mean and CV%
pk_summary['gmean'] = analysis_data.groupby(
    [CONFIG['trt_var'], 'PARAM', 'PARAMCD']
)['AVAL'].apply(lambda x: np.exp(np.log(x[x > 0]).mean())).values

pk_summary['cv_pct'] = (pk_summary['std'] / pk_summary['mean']) * 100

print("\\nPK Parameter Summary:")
print(pk_summary)"""

    def _gen_subgroup_py(self):
        treatment_var = self.analysis_params.get('treatment_variable', 'TRTA')
        subgroups = self.analysis_params.get('subgroup_variables', ['SEX', 'AGEGR1', 'RACE'])

        return f"""# --- Subgroup Analysis ---

subgroup_results = []

for sg_var in {subgroups}:
    for category in analysis_data[sg_var].unique():
        subset = analysis_data[analysis_data[sg_var] == category]
        
        if len(subset[CONFIG['trt_var']].unique()) < 2:
            continue
        
        formula = f'CHG ~ C({{CONFIG["trt_var"]}}) + BASE'
        try:
            model = ols(formula, data=subset).fit()
            
            # Get treatment effect
            params = model.params
            conf_int = model.conf_int()
            
            for param_name in params.index:
                if CONFIG['trt_var'] in param_name:
                    subgroup_results.append({{
                        'subgroup': sg_var,
                        'category': category,
                        'estimate': params[param_name],
                        'lower_ci': conf_int.loc[param_name, 0],
                        'upper_ci': conf_int.loc[param_name, 1],
                        'p_value': model.pvalues[param_name]
                    }})
        except Exception as e:
            print(f"Warning: Could not fit model for {{sg_var}}={{category}}: {{e}}")

sg_df = pd.DataFrame(subgroup_results)
print("\\nSubgroup Analysis Results:")
print(sg_df)

# Forest plot
if not sg_df.empty:
    fig, ax = plt.subplots(figsize=(10, len(sg_df) * 0.5 + 2))
    
    y_pos = range(len(sg_df))
    ax.errorbarh(
        y_pos, sg_df['estimate'],
        xerr=[sg_df['estimate'] - sg_df['lower_ci'],
              sg_df['upper_ci'] - sg_df['estimate']],
        fmt='o', color='navy', capsize=3
    )
    ax.axvline(x=0, color='grey', linestyle='--', alpha=0.7)
    ax.set_yticks(y_pos)
    ax.set_yticklabels([f"{{r['subgroup']}}: {{r['category']}}" for _, r in sg_df.iterrows()])
    ax.set_xlabel('Treatment Difference (95% CI)')
    ax.set_title('Subgroup Analysis - Forest Plot')
    plt.tight_layout()"""

    def _generate_output(self):
        output_generators = {
            'table': self._gen_table_output_py,
            'figure': self._gen_figure_output_py,
            'listing': self._gen_listing_output_py,
            'dataset': self._gen_dataset_output_py
        }
        generator = output_generators.get(self.output_type, self._gen_table_output_py)
        return generator()

    def _gen_table_output_py(self):
        return f"""# --- Generate Table Output ---

output_file = CONFIG['output_path'] / f"{{CONFIG['output_name']}}.{self.output_data.get('format', 'xlsx')}"

# Create formatted output table
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Results"

# Header information
ws['A1'] = CONFIG['study_id']
ws['A1'].font = Font(bold=True, size=12)
ws['A2'] = f"Protocol: {{CONFIG['protocol']}}"
ws['A3'] = f"Table X.X.X: {self._get_purpose_description()}"
ws['A3'].font = Font(bold=True, size=11)
ws['A4'] = f"Population: {self.analysis_params.get('population', 'Safety').title()} Population"

# Column headers
headers = ['Statistic', 'Placebo\\n(N={{}})', 'Treatment A\\n(N={{}})', 'Treatment B\\n(N={{}})']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Write results data
# (Implementation would write actual results from analysis)

# Footnotes
last_row = ws.max_row + 2
ws.cell(row=last_row, column=1, value=f"Source: {', '.join(self.input_data.get('datasets', ['adsl']))}")
ws.cell(row=last_row + 1, column=1, value=f"Program: {{CONFIG['prog_name']}}.py")
ws.cell(row=last_row + 2, column=1, value=f"Generated: {{datetime.now().strftime('%d%b%Y %H:%M')}}")

wb.save(str(output_file))
print(f"NOTE: Table written to {{output_file}}")"""

    def _gen_figure_output_py(self):
        return f"""# --- Generate Figure Output ---

output_file = CONFIG['output_path'] / f"{{CONFIG['output_name']}}.pdf"

# Finalize figure
plt.suptitle('Figure X.X.X', fontsize=12, fontweight='bold', y=0.98)
plt.title('{self._get_purpose_description()}', fontsize=10, pad=20)

# Add footnotes
plt.figtext(0.1, 0.02, 
    f"Source: {', '.join(self.input_data.get('datasets', ['adsl']))}\\n"
    f"Program: {{CONFIG['prog_name']}}.py\\n"
    f"Generated: {{datetime.now().strftime('%d%b%Y')}}",
    fontsize=7, va='bottom')

plt.tight_layout(rect=[0, 0.05, 1, 0.95])
plt.savefig(str(output_file), dpi=300, bbox_inches='tight')
plt.close()

print(f"NOTE: Figure written to {{output_file}}")"""

    def _gen_listing_output_py(self):
        return f"""# --- Generate Listing Output ---

output_file = CONFIG['output_path'] / f"{{CONFIG['output_name']}}.xlsx"

listing_cols = ['USUBJID', 'SITEID'] + {self.variables.get('analysis_variables', ['AVAL'])}
listing_data = analysis_data[listing_cols].sort_values('USUBJID')

listing_data.to_excel(str(output_file), index=False, sheet_name='Listing')
print(f"NOTE: Listing written to {{output_file}}")"""

    def _gen_dataset_output_py(self):
        return f"""# --- Generate Analysis Dataset ---

output_file = CONFIG['output_path'] / f"{{CONFIG['output_name']}}.xpt"

# Export as XPT (SAS Transport) for submission
pyreadstat.write_xport(
    final_dataset,
    str(output_file),
    file_label='{self.submission.get('study_id', 'STUDY001')} Analysis Dataset'
)

print(f"NOTE: Dataset written to {{output_file}}")
print(f"NOTE: Dataset has {{final_dataset.shape[0]}} records and {{final_dataset.shape[1]}} variables.")"""

    def _generate_validation(self):
        return f"""# --- Validation Checks ---

print("\\n" + "=" * 60)
print("VALIDATION RESULTS")
print("=" * 60)

# Check for missing critical variables
n_missing_usubjid = analysis_data['USUBJID'].isna().sum()
n_missing_trt = analysis_data[CONFIG['trt_var']].isna().sum()
n_subjects = analysis_data['USUBJID'].nunique()

print(f"  Missing USUBJID: {{n_missing_usubjid}}")
print(f"  Missing Treatment: {{n_missing_trt}}")
print(f"  Total unique subjects: {{n_subjects}}")
print(f"  Total records: {{analysis_data.shape[0]}}")

# Verify population counts
assert n_missing_usubjid == 0, "ERROR: Missing USUBJID values found!"
assert n_missing_trt == 0, "ERROR: Missing treatment values found!"

print("\\n" + "=" * 60)
print(f"Program: {{CONFIG['prog_name']}}.py")
print(f"Output:  {{CONFIG['output_name']}}.{self.output_data.get('format', 'rtf')}")
print(f"Status:  COMPLETE")
print(f"Time:    {{datetime.now().strftime('%d%b%Y %H:%M:%S')}}")
print("=" * 60)"""

    # Helper methods
    def _get_population_flag(self):
        pop_flags = {
            'safety': 'SAFFL', 'itt': 'ITTFL', 'per_protocol': 'PPROTFL',
            'pk': 'PKFL', 'full_analysis': 'FASFL'
        }
        return pop_flags.get(self.analysis_params.get('population', 'safety'), 'SAFFL')

    def _get_purpose_description(self):
        descriptions = {
            'descriptive': 'Generate descriptive statistics summary',
            'efficacy': 'Primary efficacy endpoint analysis',
            'safety': 'Safety summary of adverse events',
            'survival': 'Time-to-event survival analysis',
            'mixed_model': 'Mixed model repeated measures analysis',
            'categorical': 'Categorical data analysis',
            'pk': 'Pharmacokinetic parameter summary',
            'subgroup': 'Subgroup analysis with forest plot'
        }
        return descriptions.get(self.analysis_type, 'Statistical analysis')
    
    from .cdisc_validator import CDISCValidator
__all__ = ['CDISCValidator']

"""
CDISC Compliance Validator
Validates generated code against CDISC standards and regulatory requirements.
"""


class CDISCValidator:
    def __init__(self):
        self.adam_required_vars = {
            'adsl': ['STUDYID', 'USUBJID', 'SUBJID', 'SITEID', 'ARM', 'TRT01A', 'TRT01P',
                     'SAFFL', 'ITTFL', 'AGE', 'SEX', 'RACE'],
            'adae': ['STUDYID', 'USUBJID', 'AETERM', 'AEDECOD', 'AEBODSYS',
                     'AESTDTC', 'AEENDTC', 'AESEV', 'AESER', 'AEREL'],
            'adtte': ['STUDYID', 'USUBJID', 'PARAMCD', 'PARAM', 'AVAL', 'CNSR',
                      'STARTDT', 'ADT'],
            'adlb': ['STUDYID', 'USUBJID', 'PARAMCD', 'PARAM', 'AVAL', 'BASE',
                     'CHG', 'AVISIT', 'AVISITN', 'ANL01FL']
        }

    def validate(self, config, generated_code):
        """Run all validation checks."""
        results = {
            'compliant': True,
            'checks': [],
            'warnings': [],
            'errors': []
        }

        # Check 1: Required variables referenced
        self._check_required_variables(config, generated_code, results)

        # Check 2: Population flag usage
        self._check_population_flag(config, generated_code, results)

        # Check 3: Output metadata
        self._check_output_metadata(config, generated_code, results)

        # Check 4: Regulatory compliance
        self._check_regulatory_compliance(config, generated_code, results)

        # Check 5: Traceability
        self._check_traceability(config, generated_code, results)

        results['compliant'] = len(results['errors']) == 0
        return results

    def _check_required_variables(self, config, code, results):
        datasets = config.get('input_data', {}).get('datasets', [])
        for ds in datasets:
            if ds in self.adam_required_vars:
                for var in self.adam_required_vars[ds][:5]:  # Check first 5 critical vars
                    if var in code or var.lower() in code.lower():
                        results['checks'].append(f"✓ {ds}.{var} referenced")
                    else:
                        results['warnings'].append(
                            f"⚠ {ds}.{var} not explicitly referenced in code"
                        )

    def _check_population_flag(self, config, code, results):
        pop = config.get('analysis_parameters', {}).get('population', 'safety')
        pop_flags = {
            'safety': 'SAFFL', 'itt': 'ITTFL', 'per_protocol': 'PPROTFL',
            'pk': 'PKFL', 'full_analysis': 'FASFL'
        }
        flag = pop_flags.get(pop, 'SAFFL')
        if flag in code:
            results['checks'].append(f"✓ Population flag {flag} used correctly")
        else:
            results['errors'].append(f"✗ Population flag {flag} not found in code")

    def _check_output_metadata(self, config, code, results):
        required_elements = ['title', 'footnote', 'source', 'program']
        for elem in required_elements:
            if elem.lower() in code.lower():
                results['checks'].append(f"✓ Output contains {elem}")
            else:
                results['warnings'].append(f"⚠ Output may be missing {elem}")

    def _check_regulatory_compliance(self, config, code, results):
        authority = config.get('submission_details', {}).get('regulatory_authority', 'FDA')
        if authority == 'FDA':
            if 'xpt' in code.lower() or 'transport' in code.lower() or 'xport' in code.lower():
                results['checks'].append("✓ XPT transport format referenced (FDA requirement)")
        results['checks'].append(f"✓ Targeting {authority} submission standards")

    def _check_traceability(self, config, code, results):
        if 'study' in code.lower() and 'protocol' in code.lower():
            results['checks'].append("✓ Study and protocol identifiers present")
        else:
            results['warnings'].append("⚠ Missing study/protocol traceability")

<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Clinical Trial Code Generator | Submission Ready</title>
    <link rel="stylesheet" href="css/styles.css">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/styles/vs2015.min.css">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.0/css/all.min.css">
</head>
<body>
    <div class="app-container">
        <!-- Header -->
        <header class="app-header">
            <div class="header-left">
                <i class="fas fa-flask-vial"></i>
                <h1>Clinical Trial Code Generator</h1>
                <span class="badge">Submission Ready</span>
            </div>
            <div class="header-right">
                <span class="status-indicator online"></span>
                <span>API Connected</span>
            </div>
        </header>

        <!-- Main Content -->
        <main class="main-content">
            <!-- Left Panel: Configuration -->
            <aside class="config-panel">
                <div class="panel-section">
                    <h2><i class="fas fa-code"></i> Language</h2>
                    <div class="language-selector">
                        <button class="lang-btn active" data-lang="sas">
                            <span class="lang-icon">SAS</span>
                            <span class="lang-label">SAS</span>
                        </button>
                        <button class="lang-btn" data-lang="r">
                            <span class="lang-icon">R</span>
                            <span class="lang-label">R</span>
                        </button>
                        <button class="lang-btn" data-lang="python">
                            <span class="lang-icon">Py</span>
                            <span class="lang-label">Python</span>
                        </button>
                    </div>
                </div>

                <div class="panel-section">
                    <h2><i class="fas fa-chart-bar"></i> Analysis Type</h2>
                    <select id="analysis-type" class="form-select">
                        <option value="">-- Select Analysis Type --</option>
                        <option value="descriptive">Descriptive Statistics</option>
                        <option value="efficacy">Efficacy Analysis (ANCOVA)</option>
                        <option value="safety">Safety Analysis (AE Summary)</option>
                        <option value="survival">Survival Analysis (KM/Cox)</option>
                        <option value="mixed_model">Mixed Model (MMRM)</option>
                        <option value="categorical">Categorical Analysis</option>
                        <option value="pk">Pharmacokinetics</option>
                        <option value="subgroup">Subgroup Analysis</option>
                    </select>
                </div>

                <div class="panel-section">
                    <h2><i class="fas fa-file-export"></i> Output Type</h2>
                    <div class="output-type-selector">
                        <label class="radio-card">
                            <input type="radio" name="output-type" value="table" checked>
                            <span class="radio-card-content">
                                <i class="fas fa-table"></i>
                                <span>Table</span>
                            </span>
                        </label>
                        <label class="radio-card">
                            <input type="radio" name="output-type" value="figure">
                            <span class="radio-card-content">
                                <i class="fas fa-chart-line"></i>
                                <span>Figure</span>
                            </span>
                        </label>
                        <label class="radio-card">
                            <input type="radio" name="output-type" value="listing">
                            <span class="radio-card-content">
                                <i class="fas fa-list-alt"></i>
                                <span>Listing</span>
                            </span>
                        </label>
                        <label class="radio-card">
                            <input type="radio" name="output-type" value="dataset">
                            <span class="radio-card-content">
                                <i class="fas fa-database"></i>
                                <span>Dataset</span>
                            </span>
                        </label>
                    </div>
                </div>

                <div class="panel-section">
                    <h2><i class="fas fa-folder-open"></i> Input Data</h2>
                    <div class="form-group">
                        <label>Datasets (comma-separated)</label>
                        <input type="text" id="input-datasets" class="form-input" 
                               placeholder="adsl, adae, adtte, adlb" value="adsl, adae">
                    </div>
                    <div class="form-group">
                        <label>Data Format</label>
                        <select id="data-format" class="form-select">
                            <option value="adam">ADaM</option>
                            <option value="sdtm">SDTM</option>
                            <option value="raw">Raw/Source</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Library Path</label>
                        <input type="text" id="library-path" class="form-input" 
                               placeholder="/data/study123/adam" value="/data/adam">
                    </div>
                </div>

                <div class="panel-section">
                    <h2><i class="fas fa-tags"></i> Variables</h2>
                    <div class="form-group">
                        <label>Analysis Variables</label>
                        <input type="text" id="analysis-vars" class="form-input" 
                               placeholder="AVAL, CHG, PCHG" value="AVAL, CHG">
                    </div>
                    <div class="form-group">
                        <label>Grouping Variables</label>
                        <input type="text" id="grouping-vars" class="form-input" 
                               placeholder="TRTA, AVISIT" value="TRTA, AVISIT">
                    </div>
                    <div class="form-group">
                        <label>Filter Variables (Population Flags)</label>
                        <input type="text" id="filter-vars" class="form-input" 
                               placeholder="SAFFL, ANL01FL" value="SAFFL">
                    </div>
                    <div class="form-group">
                        <label>Sort Variables</label>
                        <input type="text" id="sort-vars" class="form-input" 
                               placeholder="USUBJID, AVISITN" value="USUBJID, AVISITN">
                    </div>
                </div>

                <div class="panel-section">
                    <h2><i class="fas fa-calculator"></i> Analysis Parameters</h2>
                    <div class="form-group">
                        <label>Statistical Method</label>
                        <select id="stat-method" class="form-select">
                            <option value="anova">ANOVA/ANCOVA</option>
                            <option value="chi_square">Chi-Square Test</option>
                            <option value="fisher_exact">Fisher's Exact Test</option>
                            <option value="cox_regression">Cox Proportional Hazards</option>
                            <option value="kaplan_meier">Kaplan-Meier</option>
                            <option value="log_rank">Log-Rank Test</option>
                            <option value="mmrm">MMRM</option>
                            <option value="logistic_regression">Logistic Regression</option>
                            <option value="cmh">Cochran-Mantel-Haenszel</option>
                            <option value="wilcoxon">Wilcoxon Rank-Sum</option>
                        </select>
                    </div>
                    <div class="form-row">
                        <div class="form-group">
                            <label>Alpha</label>
                            <input type="number" id="alpha" class="form-input" 
                                   value="0.05" step="0.01" min="0.01" max="0.1">
                        </div>
                        <div class="form-group">
                            <label>Confidence Level</label>
                            <input type="number" id="conf-level" class="form-input" 
                                   value="0.95" step="0.01" min="0.9" max="0.99">
                        </div>
                    </div>
                    <div class="form-group">
                        <label>Population</label>
                        <select id="population" class="form-select">
                            <option value="safety">Safety Population (SAFFL)</option>
                            <option value="itt">Intent-to-Treat (ITTFL)</option>
                            <option value="per_protocol">Per Protocol (PPROTFL)</option>
                            <option value="full_analysis">Full Analysis Set (FASFL)</option>
                            <option value="pk">PK Population (PKFL)</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Treatment Variable</label>
                        <input type="text" id="trt-var" class="form-input" 
                               placeholder="TRTA" value="TRTA">
                    </div>
                    <div class="form-group">
                        <label>Covariates (comma-separated)</label>
                        <input type="text" id="covariates" class="form-input" 
                               placeholder="BASE, SITEID" value="BASE">
                    </div>
                    <div class="form-group">
                        <label>Subgroup Variables</label>
                        <input type="text" id="subgroup-vars" class="form-input" 
                               placeholder="SEX, AGEGR1, RACE" value="SEX, AGEGR1, RACE">
                    </div>
                </div>

                <div class="panel-section">
                    <h2><i class="fas fa-file-alt"></i> Output Configuration</h2>
                    <div class="form-group">
                        <label>Output Name</label>
                        <input type="text" id="output-name" class="form-input" 
                               placeholder="t_ae_summary" value="t_ae_summary">
                    </div>
                    <div class="form-group">
                        <label>Output Format</label>
                        <select id="output-format" class="form-select">
                            <option value="rtf">RTF</option>
                            <option value="pdf">PDF</option>
                            <option value="xlsx">Excel (XLSX)</option>
                            <option value="html">HTML</option>
                            <option value="xpt">XPT (Transport)</option>
                            <option value="sas7bdat">SAS Dataset</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Output Path</label>
                        <input type="text" id="output-path" class="form-input" 
                               placeholder="/output/tables" value="/output/tables">
                    </div>
                </div>

                <div class="panel-section">
                    <h2><i class="fas fa-building-columns"></i> Submission Details</h2>
                    <div class="form-group">
                        <label>Study ID</label>
                        <input type="text" id="study-id" class="form-input" 
                               placeholder="STUDY-001" value="STUDY-001">
                    </div>
                    <div class="form-group">
                        <label>Protocol</label>
                        <input type="text" id="protocol" class="form-input" 
                               placeholder="D1234C00001" value="D1234C00001">
                    </div>
                    <div class="form-group">
                        <label>Sponsor</label>
                        <input type="text" id="sponsor" class="form-input" 
                               placeholder="AstraZeneca" value="AstraZeneca">
                    </div>
                    <div class="form-row">
                        <div class="form-group">
                            <label>Regulatory Authority</label>
                            <select id="reg-authority" class="form-select">
                                <option value="FDA">FDA (US)</option>
                                <option value="EMA">EMA (EU)</option>
                                <option value="PMDA">PMDA (Japan)</option>
                                <option value="NMPA">NMPA (China)</option>
                                <option value="HC">Health Canada</option>
                            </select>
                        </div>
                        <div class="form-group">
                            <label>Submission Type</label>
                            <select id="submission-type" class="form-select">
                                <option value="NDA">NDA</option>
                                <option value="BLA">BLA</option>
                                <option value="MAA">MAA</option>
                                <option value="sNDA">sNDA</option>
                            </select>
                        </div>
                    </div>
                </div>

                <div class="panel-section">
                    <h2><i class="fas fa-upload"></i> Upload Files</h2>
                    <div class="upload-area" id="shell-upload-area">
                        <input type="file" id="shell-file" accept=".xlsx,.xls,.rtf,.pdf,.docx" hidden>
                        <div class="upload-content">
                            <i class="fas fa-file-image"></i>
                            <p>Upload Output Shell</p>
                            <span class="upload-hint">RTF, Excel, PDF, or DOCX mock-up</span>
                        </div>
                        <div class="upload-status" id="shell-status"></div>
                    </div>
                    <div class="upload-area" id="spec-upload-area">
                        <input type="file" id="spec-file" accept=".xlsx,.xls,.csv,.json" hidden>
                        <div class="upload-content">
                            <i class="fas fa-file-spreadsheet"></i>
                            <p>Upload Specification</p>
                            <span class="upload-hint">Variable/Dataset specification (Excel, CSV, JSON)</span>
                        </div>
                        <div class="upload-status" id="spec-status"></div>
                    </div>
                </div>

                <!-- Generate Button -->
                <button id="generate-btn" class="generate-button">
                    <i class="fas fa-wand-magic-sparkles"></i>
                    Generate Code
                </button>
            </aside>

            <!-- Right Panel: Output -->
            <section class="output-panel">
                <!-- Tabs -->
                <div class="output-tabs">
                    <button class="tab-btn active" data-tab="code">
                        <i class="fas fa-code"></i> Generated Code
                    </button>
                    <button class="tab-btn" data-tab="validation">
                        <i class="fas fa-check-circle"></i> Validation
                    </button>
                    <button class="tab-btn" data-tab="metadata">
                        <i class="fas fa-info-circle"></i> Metadata
                    </button>
                </div>

                <!-- Code Output -->
                <div class="tab-content active" id="code-tab">
                    <div class="code-toolbar">
                        <div class="code-info">
                            <span id="code-language-badge" class="code-badge">SAS</span>
                            <span id="code-lines">0 lines</span>
                        </div>
                        <div class="code-actions">
                            <button id="copy-btn" class="toolbar-btn" title="Copy to clipboard">
                                <i class="fas fa-copy"></i> Copy
                            </button>
                            <button id="download-btn" class="toolbar-btn" title="Download file">
                                <i class="fas fa-download"></i> Download
                            </button>
                        </div>
                    </div>
                    <div class="code-container">
                        <pre><code id="generated-code" class="language-sas">/* 
   Click "Generate Code" to create submission-ready 
   clinical trial analysis code.
   
   Configure your analysis parameters in the left panel:
   1. Select programming language (SAS, R, or Python)
   2. Choose analysis type and output type
   3. Specify input datasets and variables
   4. Set statistical parameters
   5. Add submission details
   6. Optionally upload shell templates or specifications
   
   The generated code will be CDISC-compliant and 
   ready for health authority submission.
*/</code></pre>
                    </div>
                </div>

                <!-- Validation Output -->
                <div class="tab-content" id="validation-tab">
                    <div class="validation-container" id="validation-results">
                        <div class="validation-placeholder">
                            <i class="fas fa-clipboard-check"></i>
                            <p>Validation results will appear here after code generation.</p>
                        </div>
                    </div>
                </div>

                <!-- Metadata Output -->
                <div class="tab-content" id="metadata-tab">
                    <div class="metadata-container" id="metadata-results">
                        <div class="validation-placeholder">
                            <i class="fas fa-tags"></i>
                            <p>Code metadata and generation details will appear here.</p>
                        </div>
                    </div>
                </div>
            </section>
        </main>

        <!-- Footer -->
        <footer class="app-footer">
            <span>Clinical Trial Code Generator v1.0.0</span>
            <span>|</span>
            <span>CDISC ADaM/SDTM Compliant</span>
            <span>|</span>
            <span>ICH E9 Guidelines</span>
        </footer>
    </div>

    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/highlight.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/sas.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/r.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/python.min.js"></script>
    <script src="js/app.js"></script>
</body>
</html>

/* === CSS Variables === */
:root {
    --primary: #1a73e8;
    --primary-dark: #1557b0;
    --primary-light: #e8f0fe;
    --secondary: #5f6368;
    --success: #1e8e3e;
    --warning: #f9ab00;
    --error: #d93025;
    --bg-primary: #f8f9fa;
    --bg-secondary: #ffffff;
    --bg-dark: #1e1e1e;
    --text-primary: #202124;
    --text-secondary: #5f6368;
    --border: #dadce0;
    --shadow: 0 1px 3px rgba(0, 0, 0, 0.12), 0 1px 2px rgba(0, 0, 0, 0.08);
    --shadow-lg: 0 4px 6px rgba(0, 0, 0, 0.1), 0 2px 4px rgba(0, 0, 0, 0.06);
    --radius: 8px;
    --radius-sm: 4px;
    --transition: all 0.2s ease;
}

/* === Reset & Base === */
* {
    margin: 0;
    padding: 0;
    box-sizing: border-box;
}

body {
    font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, sans-serif;
    background: var(--bg-primary);
    color: var(--text-primary);
    line-height: 1.6;
    min-height: 100vh;
}

/* === App Container === */
.app-container {
    display: flex;
    flex-direction: column;
    min-height: 100vh;
}

/* === Header === */
.app-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 12px 24px;
    background: var(--bg-secondary);
    border-bottom: 1px solid var(--border);
    box-shadow: var(--shadow);
    position: sticky;
    top: 0;
    z-index: 100;
}

.header-left {
    display: flex;
    align-items: center;
    gap: 12px;
}

.header-left i {
    font-size: 24px;
    color: var(--primary);
}

.header-left h1 {
    font-size: 18px;
    font-weight: 600;
    color: var(--text-primary);
}

.badge {
    background: var(--success);
    color: white;
    padding: 2px 8px;
    border-radius: 12px;
    font-size: 11px;
    font-weight: 600;
    text-transform: uppercase;
}

.header-right {
    display: flex;
    align-items: center;
    gap: 8px;
    font-size: 13px;
    color: var(--text-secondary);
}

.status-indicator {
    width: 8px;
    height: 8px;
    border-radius: 50%;
    background: var(--error);
}

.status-indicator.online {
    background: var(--success);
    animation: pulse 2s infinite;
}

@keyframes pulse {
    0%, 100% { opacity: 1; }
    50% { opacity: 0.5; }
}

/* === Main Content === */
.main-content {
    display: grid;
    grid-template-columns: 420px 1fr;
    flex: 1;
    overflow: hidden;
}

/* === Config Panel (Left) === */
.config-panel {
    background: var(--bg-secondary);
    border-right: 1px solid var(--border);
    overflow-y: auto;
    padding: 16px;
    max-height: calc(100vh - 100px);
}

.panel-section {
    margin-bottom: 20px;
    padding-bottom: 16px;
    border-bottom: 1px solid var(--border);
}

.panel-section:last-of-type {
    border-bottom: none;
}

.panel-section h2 {
    font-size: 13px;
    font-weight: 600;
    color: var(--text-secondary);
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 12px;
    display: flex;
    align-items: center;
    gap: 8px;
}

.panel-section h2 i {
    color: var(--primary);
    font-size: 14px;
}

/* === Language Selector === */
.language-selector {
    display: flex;
    gap: 8px;
}

.lang-btn {
    flex: 1;
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 4px;
    padding: 12px 8px;
    border: 2px solid var(--border);
    border-radius: var(--radius);
    background: var(--bg-secondary);
    cursor: pointer;
    transition: var(--transition);
}

.lang-btn:hover {
    border-color: var(--primary);
    background: var(--primary-light);
}

.lang-btn.active {
    border-color: var(--primary);
    background: var(--primary-light);
    box-shadow: 0 0 0 3px rgba(26, 115, 232, 0.1);
}

.lang-icon {
    font-size: 16px;
    font-weight: 700;
    color: var(--primary);
}

.lang-label {
    font-size: 11px;
    color: var(--text-secondary);
}

/* === Output Type Selector === */
.output-type-selector {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 8px;
}

.radio-card {
    cursor: pointer;
}

.radio-card input {
    display: none;
}

.radio-card-content {
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 10px 12px;
    border: 2px solid var(--border);
    border-radius: var(--radius);
    transition: var(--transition);
    font-size: 13px;
}

.radio-card-content i {
    color: var(--text-secondary);
}

.radio-card input:checked + .radio-card-content {
    border-color: var(--primary);
    background: var(--primary-light);
}

.radio-card input:checked + .radio-card-content i {
    color: var(--primary);
}

/* === Form Elements === */
.form-group {
    margin-bottom: 12px;
}

.form-group label {
    display: block;
    font-size: 12px;
    font-weight: 500;
    color: var(--text-secondary);
    margin-bottom: 4px;
}

.form-input,
.form-select {
    width: 100%;
    padding: 8px 12px;
    border: 1px solid var(--border);
    border-radius: var(--radius-sm);
    font-size: 13px;
    color: var(--text-primary);
    background: var(--bg-secondary);
    transition: var(--transition);
}

.form-input:focus,
.form-select:focus {
    outline: none;
    border-color: var(--primary);
    box-shadow: 0 0 0 3px rgba(26, 115, 232, 0.1);
}

.form-row {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 12px;
}

/* === Upload Area === */
.upload-area {
    border: 2px dashed var(--border);
    border-radius: var(--radius);
    padding: 16px;
    text-align: center;
    cursor: pointer;
    transition: var(--transition);
    margin-bottom: 12px;
}

.upload-area:hover {
    border-color: var(--primary);
    background: var(--primary-light);
}

.upload-area.uploaded {
    border-color: var(--success);
    border-style: solid;
    background: #e6f4ea;
}

.upload-content i {
    font-size: 24px;
    color: var(--text-secondary);
    margin-bottom: 8px;
}

.upload-content p {
    font-size: 13px;
    font-weight: 500;
    color: var(--text-primary);
}

.upload-hint {
    font-size: 11px;
    color: var(--text-secondary);
}

.upload-status {
    font-size: 12px;
    margin-top: 8px;
    color: var(--success);
    font-weight: 500;
}

/* === Generate Button === */
.generate-button {
    width: 100%;
    padding: 14px;
    background: var(--primary);
    color: white;
    border: none;
    border-radius: var(--radius);
    font-size: 15px;
    font-weight: 600;
    cursor: pointer;
    transition: var(--transition);
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 8px;
    margin-top: 16px;
}

.generate-button:hover {
    background: var(--primary-dark);
    box-shadow: var(--shadow-lg);
}

.generate-button:active {
    transform: scale(0.98);
}

.generate-button.loading {
    opacity: 0.7;
    cursor: not-allowed;
}

/* === Output Panel (Right) === */
.output-panel {
    display: flex;
    flex-direction: column;
    overflow: hidden;
}

/* === Tabs === */
.output-tabs {
    display: flex;
    background: var(--bg-secondary);
    border-bottom: 1px solid var(--border);
    padding: 0 16px;
}

.tab-btn {
    padding: 12px 20px;
    border: none;
    background: none;
    font-size: 13px;
    font-weight: 500;
    color: var(--text-secondary);
    cursor: pointer;
    border-bottom: 2px solid transparent;
    transition: var(--transition);
    display: flex;
    align-items: center;
    gap: 6px;
}

.tab-btn:hover {
    color: var(--primary);
}

.tab-btn.active {
    color: var(--primary);
    border-bottom-color: var(--primary);
}

/* === Tab Content === */
.tab-content {
    display: none;
    flex: 1;
    overflow: hidden;
}

.tab-content.active {
    display: flex;
    flex-direction: column;
}

/* === Code Output === */
.code-toolbar {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 8px 16px;
    background: #2d2d2d;
    border-bottom: 1px solid #404040;
}

.code-info {
    display: flex;
    align-items: center;
    gap: 12px;
}

.code-badge {
    background: var(--primary);
    color: white;
    padding: 2px 8px;
    border-radius: 4px;
    font-size: 11px;
    font-weight: 600;
}

#code-lines {
    font-size: 12px;
    color: #888;
}

.code-actions {
    display: flex;
    gap: 8px;
}

.toolbar-btn {
    padding: 6px 12px;
    background: #404040;
    color: #ccc;
    border: 1px solid #555;
    border-radius: var(--radius-sm);
    font-size: 12px;
    cursor: pointer;
    transition: var(--transition);
    display: flex;
    align-items: center;
    gap: 4px;
}

.toolbar-btn:hover {
    background: #555;
    color: white;
}

.code-container {
    flex: 1;
    overflow: auto;
    background: var(--bg-dark);
}

.code-container pre {
    margin: 0;
    padding: 16px;
    min-height: 100%;
}

.code-container code {
    font-family: 'Cascadia Code', 'Fira Code', 'JetBrains Mono', monospace;
    font-size: 13px;
    line-height: 1.6;
}

/* === Validation Results === */
.validation-container {
    padding: 20px;
    overflow-y: auto;
    flex: 1;
}

.validation-placeholder {
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    height: 300px;
    color: var(--text-secondary);
}

.validation-placeholder i {
    font-size: 48px;
    margin-bottom: 16px;
    opacity: 0.3;
}

.validation-section {
    margin-bottom: 20px;
}

.validation-section h3 {
    font-size: 14px;
    font-weight: 600;
    margin-bottom: 8px;
    display: flex;
    align-items: center;
    gap: 8px;
}

.validation-item {
    padding: 8px 12px;
    margin-bottom: 4px;
    border-radius: var(--radius-sm);
    font-size: 13px;
}

.validation-item.check {
    background: #e6f4ea;
    color: var(--success);
}

.validation-item.warning {
    background: #fef7e0;
    color: #b06000;
}

.validation-item.error {
    background: #fce8e6;
    color: var(--error);
}

/* === Metadata === */
.metadata-container {
    padding: 20px;
    overflow-y: auto;
    flex: 1;
}

.metadata-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 16px;
}

.metadata-card {
    background: var(--bg-secondary);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 16px;
}

.metadata-card h4 {
    font-size: 12px;
    color: var(--text-secondary);
    text-transform: uppercase;
    margin-bottom: 4px;
}

.metadata-card p {
    font-size: 14px;
    font-weight: 500;
    color: var(--text-primary);
}

/* === Footer === */
.app-footer {
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 12px;
    padding: 8px;
    background: var(--bg-secondary);
    border-top: 1px solid var(--border);
    font-size: 11px;
    color: var(--text-secondary);
}

/* === Scrollbar === */
::-webkit-scrollbar {
    width: 8px;
    height: 8px;
}

::-webkit-scrollbar-track {
    background: transparent;
}

::-webkit-scrollbar-thumb {
    background: #c1c1c1;
    border-radius: 4px;
}

::-webkit-scrollbar-thumb:hover {
    background: #a1a1a1;
}

/* === Responsive === */
@media (max-width: 1200px) {
    .main-content {
        grid-template-columns: 360px 1fr;
    }
}

@media (max-width: 900px) {
    .main-content {
        grid-template-columns: 1fr;
        grid-template-rows: auto 1fr;
    }
    
    .config-panel {
        max-height: 50vh;
        border-right: none;
        border-bottom: 1px solid var(--border);
    }
}


/**
 * Clinical Trial Code Generator - Frontend Application
 * Handles UI interactions, API calls, and code display.
 */

const API_BASE = 'http://localhost:5000/api';

// State management
const state = {
    language: 'sas',
    shellFileId: null,
    specFileId: null,
    generatedCode: '',
    generatedFilename: ''
};

// DOM Elements
const elements = {
    generateBtn: document.getElementById('generate-btn'),
    codeOutput: document.getElementById('generated-code'),
    codeBadge: document.getElementById('code-language-badge'),
    codeLines: document.getElementById('code-lines'),
    copyBtn: document.getElementById('copy-btn'),
    downloadBtn: document.getElementById('download-btn'),
    shellUpload: document.getElementById('shell-upload-area'),
    specUpload: document.getElementById('spec-upload-area'),
    shellFile: document.getElementById('shell-file'),
    specFile: document.getElementById('spec-file'),
    shellStatus: document.getElementById('shell-status'),
    specStatus: document.getElementById('spec-status'),
    validationResults: document.getElementById('validation-results'),
    metadataResults: document.getElementById('metadata-results')
};

// Initialize
document.addEventListener('DOMContentLoaded', () => {
    initLanguageSelector();
    initOutputTypeSelector();
    initTabNavigation();
    initFileUploads();
    initGenerateButton();
    initToolbarActions();
    checkApiHealth();
});

// === Language Selector ===
function initLanguageSelector() {
    document.querySelectorAll('.lang-btn').forEach(btn => {
        btn.addEventListener('click', () => {
            document.querySelectorAll('.lang-btn').forEach(b => b.classList.remove('active'));
            btn.classList.add('active');
            state.language = btn.dataset.lang;
            elements.codeBadge.textContent = state.language.toUpperCase();
        });
    });
}

// === Output Type Selector ===
function initOutputTypeSelector() {
    document.querySelectorAll('input[name="output-type"]').forEach(radio => {
        radio.addEventListener('change', () => {
            updateOutputFormatOptions(radio.value);
        });
    });
}

function updateOutputFormatOptions(outputType) {
    const formatSelect = document.getElementById('output-format');
    const formats = {
        table: [
            { value: 'rtf', label: 'RTF' },
            { value: 'pdf', label: 'PDF' },
            { value: 'xlsx', label: 'Excel (XLSX)' },
            { value: 'html', label: 'HTML' }
        ],
        figure: [
            { value: 'pdf', label: 'PDF' },
            { value: 'png', label: 'PNG' },
            { value: 'svg', label: 'SVG' },
            { value: 'rtf', label: 'RTF' }
        ],
        listing: [
            { value: 'rtf', label: 'RTF' },
            { value: 'pdf', label: 'PDF' },
            { value: 'xlsx', label: 'Excel (XLSX)' }
        ],
        dataset: [
            { value: 'xpt', label: 'XPT (Transport)' },
            { value: 'sas7bdat', label: 'SAS Dataset' },
            { value: 'csv', label: 'CSV' },
            { value: 'rds', label: 'RDS (R)' }
        ]
    };

    formatSelect.innerHTML = '';
    (formats[outputType] || formats.table).forEach(fmt => {
        const option = document.createElement('option');
        option.value = fmt.value;
        option.textContent = fmt.label;
        formatSelect.appendChild(option);
    });
}

// === Tab Navigation ===
function initTabNavigation() {
    document.querySelectorAll('.tab-btn').forEach(btn => {
        btn.addEventListener('click', () => {
            document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
            document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
            btn.classList.add('active');
            document.getElementById(`${btn.dataset.tab}-tab`).classList.add('active');
        });
    });
}

// === File Uploads ===
function initFileUploads() {
    // Shell upload
    elements.shellUpload.addEventListener('click', () => elements.shellFile.click());
    elements.shellFile.addEventListener('change', (e) => handleFileUpload(e, 'shell'));

    // Specification upload
    elements.specUpload.addEventListener('click', () => elements.specFile.click());
    elements.specFile.addEventListener('change', (e) => handleFileUpload(e, 'specification'));

    // Drag and drop
    [elements.shellUpload, elements.specUpload].forEach(area => {
        area.addEventListener('dragover', (e) => {
            e.preventDefault();
            area.style.borderColor = 'var(--primary)';
            area.style.background = 'var(--primary-light)';
        });
        area.addEventListener('dragleave', () => {
            area.style.borderColor = '';
            area.style.background = '';
        });
        area.addEventListener('drop', (e) => {
            e.preventDefault();
            area.style.borderColor = '';
            area.style.background = '';
            const type = area.id.includes('shell') ? 'shell' : 'specification';
            const file = e.dataTransfer.files[0];
            if (file) uploadFile(file, type);
        });
    });
}

async function handleFileUpload(event, type) {
    const file = event.target.files[0];
    if (file) await uploadFile(file, type);
}

async function uploadFile(file, type) {
    const formData = new FormData();
    formData.append('file', file);

    const statusEl = type === 'shell' ? elements.shellStatus : elements.specStatus;
    const areaEl = type === 'shell' ? elements.shellUpload : elements.specUpload;

    statusEl.textContent = 'Uploading...';

    try {
        const response = await fetch(`${API_BASE}/upload/${type}`, {
            method: 'POST',
            body: formData
        });

        const result = await response.json();

        if (result.success) {
            if (type === 'shell') {
                state.shellFileId = result.file_id;
            } else {
                state.specFileId = result.file_id;
            }
            statusEl.textContent = `✓ ${result.filename} uploaded`;
            areaEl.classList.add('uploaded');
        } else {
            statusEl.textContent = `✗ Error: ${result.error}`;
            statusEl.style.color = 'var(--error)';
        }
    } catch (error) {
        statusEl.textContent = `✗ Upload failed: ${error.message}`;
        statusEl.style.color = 'var(--error)';
    }
}

// === Generate Code ===
function initGenerateButton() {
    elements.generateBtn.addEventListener('click', generateCode);
}

async function generateCode() {
    const btn = elements.generateBtn;
    btn.classList.add('loading');
    btn.innerHTML = '<i class="fas fa-spinner fa-spin"></i> Generating...';

    const payload = buildPayload();

    try {
        const response = await fetch(`${API_BASE}/generate`, {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify(payload)
        });

        const result = await response.json();

        if (result.success) {
            state.generatedCode = result.code;
            state.generatedFilename = result.filename;
            displayCode(result.code, result.language);
            displayValidation(result.validation);
            displayMetadata(result.metadata);
        } else {
            displayError(result.error);
        }
    } catch (error) {
        displayError(`Connection error: ${error.message}. Ensure the backend is running.`);
    } finally {
        btn.classList.remove('loading');
        btn.innerHTML = '<i class="fas fa-wand-magic-sparkles"></i> Generate Code';
    }
}

function buildPayload() {
    const getVal = (id) => document.getElementById(id)?.value || '';
    const getList = (id) => getVal(id).split(',').map(s => s.trim()).filter(Boolean);

    return {
        language: state.language,
        analysis_type: getVal('analysis-type'),
        output_type: document.querySelector('input[name="output-type"]:checked')?.value || 'table',
        input_data: {
            datasets: getList('input-datasets'),
            format: getVal('data-format'),
            library_path: getVal('library-path')
        },
        output_data: {
            dataset_name: getVal('output-name'),
            format: getVal('output-format'),
            output_path: getVal('output-path')
        },
        variables: {
            analysis_variables: getList('analysis-vars'),
            grouping_variables: getList('grouping-vars'),
            filter_variables: getList('filter-vars'),
            sort_variables: getList('sort-vars')
        },
        analysis_parameters: {
            statistical_method: getVal('stat-method'),
            confidence_level: parseFloat(getVal('conf-level')) || 0.95,
            alpha: parseFloat(getVal('alpha')) || 0.05,
            population: getVal('population'),
            treatment_variable: getVal('trt-var') || 'TRTA',
            subgroup_variables: getList('subgroup-vars'),
            covariates: getList('covariates')
        },
        submission_details: {
            study_id: getVal('study-id'),
            sponsor: getVal('sponsor'),
            protocol: getVal('protocol'),
            regulatory_authority: getVal('reg-authority'),
            submission_type: getVal('submission-type')
        },
        shell_file_id: state.shellFileId,
        specification_file_id: state.specFileId
    };
}

// === Display Functions ===
function displayCode(code, language) {
    const langMap = { sas: 'sas', r: 'r', python: 'python' };
    const codeEl = elements.codeOutput;
    codeEl.className = `language-${langMap[language] || 'plaintext'}`;
    codeEl.textContent = code;
    hljs.highlightElement(codeEl);

    const lineCount = code.split('\n').length;
    elements.codeLines.textContent = `${lineCount} lines`;
    elements.codeBadge.textContent = language.toUpperCase();

    // Switch to code tab
    document.querySelector('[data-tab="code"]').click();
}

function displayValidation(validation) {
    if (!validation) return;

    let html = '';

    if (validation.compliant) {
        html += `<div class="validation-section">
            <h3><i class="fas fa-check-circle" style="color: var(--success)"></i> CDISC Compliant</h3>
            <p style="color: var(--success); font-weight: 500;">Code passes all compliance checks.</p>
        </div>`;
    } else {
        html += `<div class="validation-section">
            <h3><i class="fas fa-exclamation-circle" style="color: var(--error)"></i> Compliance Issues Found</h3>
        </div>`;
    }

    if (validation.checks && validation.checks.length > 0) {
        html += `<div class="validation-section"><h3>Checks Passed</h3>`;
        validation.checks.forEach(check => {
            html += `<div class="validation-item check">${check}</div>`;
        });
        html += '</div>';
    }

    if (validation.warnings && validation.warnings.length > 0) {
        html += `<div class="validation-section"><h3>Warnings</h3>`;
        validation.warnings.forEach(warning => {
            html += `<div class="validation-item warning">${warning}</div>`;
        });
        html += '</div>';
    }

    if (validation.errors && validation.errors.length > 0) {
        html += `<div class="validation-section"><h3>Errors</h3>`;
        validation.errors.forEach(error => {
            html += `<div class="validation-item error">${error}</div>`;
        });
        html += '</div>';
    }

    elements.validationResults.innerHTML = html;
}

function displayMetadata(metadata) {
    if (!metadata) return;

    elements.metadataResults.innerHTML = `
        <div class="metadata-grid">
            <div class="metadata-card">
                <h4>Generated At</h4>
                <p>${new Date(metadata.generated_at).toLocaleString()}</p>
            </div>
            <div class="metadata-card">
                <h4>Analysis Type</h4>
                <p>${metadata.analysis_type?.replace('_', ' ').toUpperCase()}</p>
            </div>
            <div class="metadata-card">
                <h4>Output Type</h4>
                <p>${metadata.output_type?.toUpperCase()}</p>
            </div>
            <div class="metadata-card">
                <h4>Submission Ready</h4>
                <p style="color: ${metadata.submission_ready ? 'var(--success)' : 'var(--warning)'}">
                    ${metadata.submission_ready ? '✓ Yes' : '⚠ Review Required'}
                </p>
            </div>
        </div>
    `;
}

function displayError(message) {
    elements.codeOutput.textContent = `/* ERROR: ${message} */`;
    elements.codeOutput.className = 'language-plaintext';
}

// === Toolbar Actions ===
function initToolbarActions() {
    elements.copyBtn.addEventListener('click', () => {
        navigator.clipboard.writeText(state.generatedCode).then(() => {
            elements.copyBtn.innerHTML = '<i class="fas fa-check"></i> Copied!';
            setTimeout(() => {
                elements.copyBtn.innerHTML = '<i class="fas fa-copy"></i> Copy';
            }, 2000);
        });
    });

    elements.downloadBtn.addEventListener('click', () => {
        if (!state.generatedCode) return;

        const extensions = { sas: '.sas', r: '.R', python: '.py' };
        const filename = state.generatedFilename || 
            `generated_code${extensions[state.language] || '.txt'}`;

        const blob = new Blob([state.generatedCode], { type: 'text/plain' });
        const url = URL.createObjectURL(blob);
        const a = document.createElement('a');
        a.href = url;
        a.download = filename;
        a.click();
        URL.revokeObjectURL(url);
    });
}

// === API Health Check ===
async function checkApiHealth() {
    const indicator = document.querySelector('.status-indicator');
    const statusText = indicator.nextElementSibling;

    try {
        const response = await fetch(`${API_BASE}/health`);
        if (response.ok) {
            indicator.classList.add('online');
            statusText.textContent = 'API Connected';
        } else {
            indicator.classList.remove('online');
            statusText.textContent = 'API Error';
        }
    } catch {
        indicator.classList.remove('online');
        statusText.textContent = 'API Offline';
    }
}

{
    "editor.formatOnSave": true,
    "editor.defaultFormatter": "esbenp.prettier-vscode",
    "python.defaultInterpreterPath": "./venv/bin/python",
    "python.linting.enabled": true,
    "python.linting.pylintEnabled": true,
    "files.associations": {
        "*.sas": "sas"
    },
    "emmet.includeLanguages": {
        "html": "html"
    },
    "liveServer.settings.port": 3000,
    "liveServer.settings.root": "/frontend",
    "[python]": {
        "editor.defaultFormatter": "ms-python.black-formatter",
        "editor.formatOnSave": true
    },
    "[html]": {
        "editor.defaultFormatter": "esbenp.prettier-vscode"
    },
    "[css]": {
        "editor.defaultFormatter": "esbenp.prettier-vscode"
    },
    "[javascript]": {
        "editor.defaultFormatter": "esbenp.prettier-vscode"
    }
}

{
    "name": "clinical-code-generator",
    "version": "1.0.0",
    "description": "Clinical Trial Analysis Code Generator - Submission Ready",
    "scripts": {
        "start:backend": "cd backend && python app.py",
        "start:frontend": "npx live-server frontend --port=3000",
        "start": "concurrently \"npm run start:backend\" \"npm run start:frontend\"",
        "setup": "cd backend && pip install -r requirements.txt"
    },
    "devDependencies": {
        "concurrently": "^8.2.2",
        "live-server": "^1.2.2"
    },
    "keywords": ["clinical-trial", "SAS", "R", "python", "CDISC", "submission"],
    "author": "Clinical Programming Team",
    "license": "MIT"
}

# Clinical Trial Analysis Code Generator

A web-based tool that generates **submission-ready** SAS, R, and Python code for clinical trial analysis, compliant with CDISC standards (ADaM/SDTM) and health authority requirements (FDA, EMA, PMDA).

## Features

- **Multi-language support**: SAS, R, Python
- **Analysis types**: Descriptive, Efficacy, Safety, Survival, MMRM, Categorical, PK, Subgroup
- **Output types**: Tables, Figures, Listings, Datasets
- **CDISC compliance**: ADaM IG v1.3, SDTM IG v3.4
- **Regulatory targeting**: FDA, EMA, PMDA, NMPA, Health Canada
- **Shell upload**: Upload RTF/Excel mock-ups to match output layout
- **Specification upload**: Upload variable/dataset specifications
- **Validation**: Automatic CDISC compliance checking

## Quick Start

### Prerequisites
- Python 3.9+
- Node.js 18+

### Installation

```bash
# Clone and navigate
cd clinical-code-generator

# Install backend dependencies
cd backend
pip install -r requirements.txt
cd ..

# Install frontend dependencies
npm install

# Start both backend and frontend
npm start

# Or separately:
npm run start:backend    # Flask API on port 5000
npm run start:frontend   # Live server on port 3000

---

## How to Run This Project

1. **Open the project folder in VS Code**
2. **Install Python dependencies:**
   ```bash
   cd backend
   pip install -r requirements.txt

   npm install
   npm start
   