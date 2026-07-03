"""
================================================================================
PROGRAM NAME  : t_ae_summary.py
STUDY         : STUDY-001
PROTOCOL      : D1234C00001
SPONSOR       : AstraZeneca
PURPOSE       : Time-to-event survival analysis
INPUT         : adsl, adae
OUTPUT        : t_ae_summary.rtf
ANALYSIS TYPE : Survival
OUTPUT TYPE   : Table
POPULATION    : Safety

AUTHOR        : [Programmer Name]
DATE CREATED  : 22-Jun-2026
DATE MODIFIED : 22-Jun-2026

REGULATORY    : FDA
SUBMISSION    : NDA

VALIDATION    : Independent double programming required
CDISC VERSION : ADaM IG v1.3 / SDTM IG v3.4
================================================================================
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")
from lifelines import KaplanMeierFitter, CoxPHFitter
from lifelines.statistics import logrank_test
import matplotlib.pyplot as plt
from openpyxl import Workbook
from docx import Document

# --- Program Configuration ---
CONFIG = {
    'study_id': 'STUDY-001',
    'protocol': 'D1234C00001',
    'prog_name': 't_ae_summary',
    'output_name': 't_ae_summary',
    'pop_flag': 'SAFFL',
    'trt_var': 'TRTA',
    'alpha': 0.05,
    'conf_level': 0.95,
    'adam_path': Path(r'/data/adam'),
    'output_path': Path(r'/output/tables'),
}

# Treatment ordering
TRT_ORDER = ['Placebo', 'Treatment A', 'Treatment B', 'Total']
TRT_MAP = {'Placebo': 1, 'Treatment A': 2, 'Treatment B': 3, 'Total': 99}

# Ensure output directory exists
CONFIG['output_path'].mkdir(parents=True, exist_ok=True)

# --- Read Input Datasets ---
import pyreadstat

# Read ADSL
adsl_path = CONFIG['adam_path'] / 'adsl.sas7bdat'
adsl_raw, meta = pyreadstat.read_sas7bdat(str(adsl_path))
adsl = adsl_raw[(df["SAFFL"] == "Y")].copy()
adsl[CONFIG['trt_var']] = pd.Categorical(
    adsl[CONFIG['trt_var']], categories=TRT_ORDER[:-1], ordered=True
)
print(f"NOTE: ADSL has {adsl.shape[0]} records after filtering.")
# Read ADAE
adae_path = CONFIG['adam_path'] / 'adae.sas7bdat'
adae_raw, meta = pyreadstat.read_sas7bdat(str(adae_path))
adae = adae_raw[(df["SAFFL"] == "Y")].copy()
adae[CONFIG['trt_var']] = pd.Categorical(
    adae[CONFIG['trt_var']], categories=TRT_ORDER[:-1], ordered=True
)
print(f"NOTE: ADAE has {adae.shape[0]} records after filtering.")

# --- Get Population Counts (Big N) ---
big_n = adsl.drop_duplicates(subset=['USUBJID']).groupby(
    CONFIG['trt_var']
)['USUBJID'].nunique().to_dict()

print(f"NOTE: Population counts: {big_n}")

# --- Data Preparation ---
analysis_data = adsl.copy()

# Add treatment numeric ordering
analysis_data['TRTN'] = analysis_data[CONFIG['trt_var']].map(TRT_MAP)

# Sort data
analysis_data = analysis_data.sort_values(
    ['TRTN'] + ['USUBJID', 'AVISITN']
).reset_index(drop=True)

# Create version with Total group
analysis_total = analysis_data.copy()
analysis_total[CONFIG['trt_var']] = 'Total'
analysis_total['TRTN'] = 99

analysis_with_total = pd.concat(
    [analysis_data, analysis_total], ignore_index=True
)

print(f"NOTE: Analysis dataset has {analysis_with_total.shape[0]} records (including Total).")

# --- Survival Analysis ---

# Kaplan-Meier Estimation
kmf = KaplanMeierFitter()

fig, ax = plt.subplots(figsize=(10, 7))
km_results = {}

for trt in analysis_data[CONFIG['trt_var']].unique():
    subset = analysis_data[analysis_data[CONFIG['trt_var']] == trt]
    kmf.fit(
        durations=subset['AVAL'],
        event_observed=(1 - subset['CNSR']),
        label=trt
    )
    kmf.plot_survival_function(ax=ax, ci_show=True)
    
    km_results[trt] = {
        'median': kmf.median_survival_time_,
        'survival_6mo': kmf.predict(180),
        'survival_12mo': kmf.predict(365)
    }

ax.set_xlabel('Time (Days)')
ax.set_ylabel('Survival Probability')
ax.set_title('Kaplan-Meier Survival Curve')
ax.legend(loc='lower left')

# Log-rank test
groups = analysis_data[CONFIG['trt_var']].unique()
if len(groups) == 2:
    g1 = analysis_data[analysis_data[CONFIG['trt_var']] == groups[0]]
    g2 = analysis_data[analysis_data[CONFIG['trt_var']] == groups[1]]
    lr_result = logrank_test(
        g1['AVAL'], g2['AVAL'],
        event_observed_A=(1 - g1['CNSR']),
        event_observed_B=(1 - g2['CNSR'])
    )
    print(f"\nLog-Rank Test p-value: {lr_result.p_value:.4f}")

# Cox Proportional Hazards
cph = CoxPHFitter()
cox_data = analysis_data[['AVAL', 'CNSR', CONFIG['trt_var']]].copy()
cox_data['event'] = 1 - cox_data['CNSR']
cox_data = pd.get_dummies(cox_data, columns=[CONFIG['trt_var']], drop_first=True)

cph.fit(cox_data, duration_col='AVAL', event_col='event')
print("\nCox PH Model:")
cph.print_summary()

# --- Generate Table Output ---

output_file = CONFIG['output_path'] / f"{CONFIG['output_name']}.rtf"

# Create formatted output table
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Results"

# Header information
ws['A1'] = CONFIG['study_id']
ws['A1'].font = Font(bold=True, size=12)
ws['A2'] = f"Protocol: {CONFIG['protocol']}"
ws['A3'] = f"Table X.X.X: Time-to-event survival analysis"
ws['A3'].font = Font(bold=True, size=11)
ws['A4'] = f"Population: Safety Population"

# Column headers
headers = ['Statistic', 'Placebo\n(N={})', 'Treatment A\n(N={})', 'Treatment B\n(N={})']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Write results data
# (Implementation would write actual results from analysis)

# Footnotes
last_row = ws.max_row + 2
ws.cell(row=last_row, column=1, value=f"Source: adsl, adae")
ws.cell(row=last_row + 1, column=1, value=f"Program: {CONFIG['prog_name']}.py")
ws.cell(row=last_row + 2, column=1, value=f"Generated: {datetime.now().strftime('%d%b%Y %H:%M')}")

wb.save(str(output_file))
print(f"NOTE: Table written to {output_file}")

# --- Validation Checks ---

print("\n" + "=" * 60)
print("VALIDATION RESULTS")
print("=" * 60)

# Check for missing critical variables
n_missing_usubjid = analysis_data['USUBJID'].isna().sum()
n_missing_trt = analysis_data[CONFIG['trt_var']].isna().sum()
n_subjects = analysis_data['USUBJID'].nunique()

print(f"  Missing USUBJID: {n_missing_usubjid}")
print(f"  Missing Treatment: {n_missing_trt}")
print(f"  Total unique subjects: {n_subjects}")
print(f"  Total records: {analysis_data.shape[0]}")

# Verify population counts
assert n_missing_usubjid == 0, "ERROR: Missing USUBJID values found!"
assert n_missing_trt == 0, "ERROR: Missing treatment values found!"

print("\n" + "=" * 60)
print(f"Program: {CONFIG['prog_name']}.py")
print(f"Output:  {CONFIG['output_name']}.rtf")
print(f"Status:  COMPLETE")
print(f"Time:    {datetime.now().strftime('%d%b%Y %H:%M:%S')}")
print("=" * 60)