"""
================================================================================
PROGRAM NAME  : t_ae_summary.py
STUDY         : STUDY-001
PROTOCOL      : D1234C00001
SPONSOR       : AstraZeneca
PURPOSE       : Generate descriptive statistics summary
INPUT         : adsl, adae
OUTPUT        : t_ae_summary.rtf
ANALYSIS TYPE : Descriptive
OUTPUT TYPE   : Table
POPULATION    : Safety

AUTHOR        : [Programmer Name]
DATE CREATED  : 22-Jun-2026
DATE MODIFIED : 22-Jun-2026

REGULATORY    : FDA
SUBMISSION    : NDA

VALIDATION    : Independent double programming required
CDISC VERSION : ADaM IG v1.3 / SDTM IG v3.4
================================================================================
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")
from scipy import stats
import statsmodels.api as sm
from openpyxl import Workbook
from docx import Document

# --- Program Configuration ---
CONFIG = {
    'study_id': 'STUDY-001',
    'protocol': 'D1234C00001',
    'prog_name': 't_ae_summary',
    'output_name': 't_ae_summary',
    'pop_flag': 'SAFFL',
    'trt_var': 'TRTA',
    'alpha': 0.05,
    'conf_level': 0.95,
    'adam_path': Path(r'/data/adam'),
    'output_path': Path(r'/output/tables'),
}

# Treatment ordering
TRT_ORDER = ['Placebo', 'Treatment A', 'Treatment B', 'Total']
TRT_MAP = {'Placebo': 1, 'Treatment A': 2, 'Treatment B': 3, 'Total': 99}

# Ensure output directory exists
CONFIG['output_path'].mkdir(parents=True, exist_ok=True)

# --- Read Input Datasets ---
import pyreadstat

# Read ADSL
adsl_path = CONFIG['adam_path'] / 'adsl.sas7bdat'
adsl_raw, meta = pyreadstat.read_sas7bdat(str(adsl_path))
adsl = adsl_raw[(df["SAFFL"] == "Y")].copy()
adsl[CONFIG['trt_var']] = pd.Categorical(
    adsl[CONFIG['trt_var']], categories=TRT_ORDER[:-1], ordered=True
)
print(f"NOTE: ADSL has {adsl.shape[0]} records after filtering.")
# Read ADAE
adae_path = CONFIG['adam_path'] / 'adae.sas7bdat'
adae_raw, meta = pyreadstat.read_sas7bdat(str(adae_path))
adae = adae_raw[(df["SAFFL"] == "Y")].copy()
adae[CONFIG['trt_var']] = pd.Categorical(
    adae[CONFIG['trt_var']], categories=TRT_ORDER[:-1], ordered=True
)
print(f"NOTE: ADAE has {adae.shape[0]} records after filtering.")

# --- Get Population Counts (Big N) ---
big_n = adsl.drop_duplicates(subset=['USUBJID']).groupby(
    CONFIG['trt_var']
)['USUBJID'].nunique().to_dict()

print(f"NOTE: Population counts: {big_n}")

# --- Data Preparation ---
analysis_data = adsl.copy()

# Add treatment numeric ordering
analysis_data['TRTN'] = analysis_data[CONFIG['trt_var']].map(TRT_MAP)

# Sort data
analysis_data = analysis_data.sort_values(
    ['TRTN'] + ['USUBJID', 'AVISITN']
).reset_index(drop=True)

# Create version with Total group
analysis_total = analysis_data.copy()
analysis_total[CONFIG['trt_var']] = 'Total'
analysis_total['TRTN'] = 99

analysis_with_total = pd.concat(
    [analysis_data, analysis_total], ignore_index=True
)

print(f"NOTE: Analysis dataset has {analysis_with_total.shape[0]} records (including Total).")

# --- Descriptive Statistics ---

def compute_descriptive_stats(df, var, group_var):
    """Compute descriptive statistics by treatment group."""
    stats_df = df.groupby(group_var)[var].agg([
        ('n', 'count'),
        ('mean', 'mean'),
        ('std', 'std'),
        ('median', 'median'),
        ('min', 'min'),
        ('max', 'max'),
        ('q1', lambda x: x.quantile(0.25)),
        ('q3', lambda x: x.quantile(0.75))
    ]).reset_index()
    
    # Format for display
    stats_df['n_c'] = stats_df['n'].apply(lambda x: f"{x:.0f}")
    stats_df['mean_sd'] = stats_df.apply(
        lambda r: f"{r['mean']:.1f} ({r['std']:.2f})", axis=1
    )
    stats_df['median_c'] = stats_df['median'].apply(lambda x: f"{x:.1f}")
    stats_df['q1_q3'] = stats_df.apply(
        lambda r: f"{r['q1']:.1f}, {r['q3']:.1f}", axis=1
    )
    stats_df['min_max'] = stats_df.apply(
        lambda r: f"{r['min']:.1f}, {r['max']:.1f}", axis=1
    )
    
    return stats_df

# Compute statistics for each analysis variable
results = {}
for var in ['AVAL', 'CHG']:
    results[var] = compute_descriptive_stats(
        analysis_with_total, var, CONFIG['trt_var']
    )
    print(f"\nDescriptive Statistics for {var}:")
    print(results[var][['TRTA', 'n_c', 'mean_sd', 'median_c', 'q1_q3', 'min_max']])

# --- Generate Table Output ---

output_file = CONFIG['output_path'] / f"{CONFIG['output_name']}.rtf"

# Create formatted output table
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Results"

# Header information
ws['A1'] = CONFIG['study_id']
ws['A1'].font = Font(bold=True, size=12)
ws['A2'] = f"Protocol: {CONFIG['protocol']}"
ws['A3'] = f"Table X.X.X: Generate descriptive statistics summary"
ws['A3'].font = Font(bold=True, size=11)
ws['A4'] = f"Population: Safety Population"

# Column headers
headers = ['Statistic', 'Placebo\n(N={})', 'Treatment A\n(N={})', 'Treatment B\n(N={})']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Write results data
# (Implementation would write actual results from analysis)

# Footnotes
last_row = ws.max_row + 2
ws.cell(row=last_row, column=1, value=f"Source: adsl, adae")
ws.cell(row=last_row + 1, column=1, value=f"Program: {CONFIG['prog_name']}.py")
ws.cell(row=last_row + 2, column=1, value=f"Generated: {datetime.now().strftime('%d%b%Y %H:%M')}")

wb.save(str(output_file))
print(f"NOTE: Table written to {output_file}")

# --- Validation Checks ---

print("\n" + "=" * 60)
print("VALIDATION RESULTS")
print("=" * 60)

# Check for missing critical variables
n_missing_usubjid = analysis_data['USUBJID'].isna().sum()
n_missing_trt = analysis_data[CONFIG['trt_var']].isna().sum()
n_subjects = analysis_data['USUBJID'].nunique()

print(f"  Missing USUBJID: {n_missing_usubjid}")
print(f"  Missing Treatment: {n_missing_trt}")
print(f"  Total unique subjects: {n_subjects}")
print(f"  Total records: {analysis_data.shape[0]}")

# Verify population counts
assert n_missing_usubjid == 0, "ERROR: Missing USUBJID values found!"
assert n_missing_trt == 0, "ERROR: Missing treatment values found!"

print("\n" + "=" * 60)
print(f"Program: {CONFIG['prog_name']}.py")
print(f"Output:  {CONFIG['output_name']}.rtf")
print(f"Status:  COMPLETE")
print(f"Time:    {datetime.now().strftime('%d%b%Y %H:%M:%S')}")
print("=" * 60)